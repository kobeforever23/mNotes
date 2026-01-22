#!/usr/bin/env python3
"""
================================================================================
UNIFIED AI BUBBLE CORRECTION STRESS SCENARIO
================================================================================
Complete A-to-Z Implementation

ASSET CLASSES:
- EQUITY (EQ): Indices, sectors, single names with AI tier classification
- CREDIT (CR): Corporate bonds, CDS, Indices with NAIC codes
- COMMODITIES (CO): Energy, precious, industrial metals
- RATES (RT): Treasuries, swaps, curves
- FX: G10 and EM currencies
- VOLATILITY (VOL): VIX, skew, term structure

SCENARIO PARAMETERS:
- Moderate: 10-day horizon, orderly correction
- Severe: 60-day horizon, disorderly unwind with contagion

METHODOLOGY:
- Historical calibration to dot-com bust, COVID crash, 2022 tech selloff
- Cross-asset consistency validation
- Beta-based propagation with binding constraints
- Full derivation trail for every shock

================================================================================
Author: Market Risk Analytics
Version: 2.0
Date: 2026-01-22
================================================================================
"""

import json
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any, Set, Union
from enum import Enum, auto
from datetime import datetime
import re


# =============================================================================
# SECTION 1: ENUMERATIONS AND CONSTANTS
# =============================================================================

class AssetClass(Enum):
    EQUITY = "Equity"
    CREDIT = "Credit"
    COMMODITY = "Commodity"
    RATES = "Rates"
    FX = "FX"
    VOLATILITY = "Volatility"


class CreditType(Enum):
    CORPORATE = "Corporate"
    CDS = "CDS"
    INDEX = "Index"


class RatingBucket(Enum):
    IG = "IG"
    HY = "HY"
    NR = "NR"


class AITier(Enum):
    """AI Exposure Classification"""
    TIER1_DIRECT = "Tier1_Direct"      # Semiconductors, AI software, data centers
    TIER2_INDIRECT = "Tier2_Indirect"  # Cloud, hyperscalers, AI utilities
    TIER3_MINIMAL = "Tier3_Minimal"    # Traditional sectors
    DEFENSIVE = "Defensive"            # Healthcare, staples


class CommodityType(Enum):
    ENERGY = "Energy"
    PRECIOUS = "Precious"
    INDUSTRIAL = "Industrial"
    AGRICULTURAL = "Agricultural"


class ValidationSeverity(Enum):
    ERROR = "Error"
    WARNING = "Warning"
    INFO = "Info"


# =============================================================================
# SECTION 2: HISTORICAL CALIBRATION DATA
# =============================================================================

HISTORICAL_EPISODES = {
    "dot_com_bust_2000_2002": {
        "description": "Dot-com bubble burst, Mar 2000 - Oct 2002",
        "spx_drawdown": -49.1,
        "ndx_drawdown": -78.4,
        "vix_peak": 45.7,
        "ig_spread_widening": 150,
        "hy_spread_widening": 650,
        "duration_months": 31,
    },
    "gfc_2008_2009": {
        "description": "Global Financial Crisis, Oct 2007 - Mar 2009",
        "spx_drawdown": -56.8,
        "ndx_drawdown": -54.0,
        "vix_peak": 80.9,
        "ig_spread_widening": 400,
        "hy_spread_widening": 1600,
        "duration_months": 17,
    },
    "covid_crash_2020": {
        "description": "COVID-19 crash, Feb 2020 - Mar 2020",
        "spx_drawdown": -33.9,
        "ndx_drawdown": -30.1,
        "vix_peak": 82.7,
        "ig_spread_widening": 200,
        "hy_spread_widening": 700,
        "duration_days": 23,
    },
    "tech_selloff_2022": {
        "description": "2022 Tech/Growth selloff, Jan 2022 - Oct 2022",
        "spx_drawdown": -25.4,
        "ndx_drawdown": -36.4,
        "vix_peak": 36.5,
        "ig_spread_widening": 80,
        "hy_spread_widening": 280,
        "duration_months": 10,
    },
}

# Our scenario calibration target (blend of episodes)
SCENARIO_CALIBRATION = {
    "moderate": {
        "horizon_days": 10,
        "description": "Orderly AI valuation correction",
        "historical_analog": "Early 2022 selloff (first 2 weeks)",
        "spx_target": -22,
        "vix_target": 38,
        "ig_oas_target": 100,
        "hy_oas_target": 180,
    },
    "severe": {
        "horizon_days": 60,
        "description": "Disorderly AI unwind with contagion",
        "historical_analog": "Blend of dot-com (-49% over 31mo scaled to 60d) + COVID pace",
        "spx_target": -33,
        "vix_target": 58,
        "ig_oas_target": 200,
        "hy_oas_target": 420,
    },
}


# =============================================================================
# SECTION 3: NAIC CODE MAPPING (COMPREHENSIVE)
# =============================================================================

NAIC_CODE_MAPPING = {
    # =========================================================================
    # TIER 1: DIRECT AI IMPACT
    # =========================================================================
    # Semiconductor and Electronic Component Manufacturing
    "334413": {"sector": "Semiconductors", "subsector": "Semiconductor Manufacturing", "ai_tier": AITier.TIER1_DIRECT},
    "334419": {"sector": "Semiconductors", "subsector": "Other Electronic Components", "ai_tier": AITier.TIER1_DIRECT},
    "334418": {"sector": "Semiconductors", "subsector": "Printed Circuit Assembly", "ai_tier": AITier.TIER1_DIRECT},
    "334416": {"sector": "Semiconductors", "subsector": "Capacitors/Resistors", "ai_tier": AITier.TIER1_DIRECT},
    "334417": {"sector": "Semiconductors", "subsector": "Electronic Connectors", "ai_tier": AITier.TIER1_DIRECT},
    
    # Computer and Electronic Product Manufacturing
    "334111": {"sector": "Technology", "subsector": "Electronic Computers", "ai_tier": AITier.TIER1_DIRECT},
    "334112": {"sector": "Technology", "subsector": "Computer Storage Devices", "ai_tier": AITier.TIER1_DIRECT},
    "334118": {"sector": "Technology", "subsector": "Computer Terminals", "ai_tier": AITier.TIER1_DIRECT},
    
    # Software and Data Processing
    "511210": {"sector": "Software", "subsector": "Software Publishers", "ai_tier": AITier.TIER1_DIRECT},
    "518210": {"sector": "Technology", "subsector": "Data Processing & Hosting", "ai_tier": AITier.TIER1_DIRECT},
    "541511": {"sector": "Technology", "subsector": "Custom Computer Programming", "ai_tier": AITier.TIER1_DIRECT},
    "541512": {"sector": "Technology", "subsector": "Computer Systems Design", "ai_tier": AITier.TIER1_DIRECT},
    
    # =========================================================================
    # TIER 2: INDIRECT AI IMPACT
    # =========================================================================
    # Internet and Telecommunications
    "518111": {"sector": "Technology", "subsector": "Internet Service Providers", "ai_tier": AITier.TIER2_INDIRECT},
    "517311": {"sector": "Telecom", "subsector": "Wired Telecommunications", "ai_tier": AITier.TIER2_INDIRECT},
    "517312": {"sector": "Telecom", "subsector": "Wireless Telecommunications", "ai_tier": AITier.TIER2_INDIRECT},
    "517410": {"sector": "Telecom", "subsector": "Satellite Telecommunications", "ai_tier": AITier.TIER2_INDIRECT},
    "517919": {"sector": "Telecom", "subsector": "Other Telecommunications", "ai_tier": AITier.TIER2_INDIRECT},
    
    # Electric Power (Data Center Power)
    "221111": {"sector": "Utilities", "subsector": "Hydroelectric Power", "ai_tier": AITier.TIER2_INDIRECT},
    "221112": {"sector": "Utilities", "subsector": "Fossil Fuel Power", "ai_tier": AITier.TIER3_MINIMAL},
    "221113": {"sector": "Utilities", "subsector": "Nuclear Power", "ai_tier": AITier.TIER2_INDIRECT},
    "221114": {"sector": "Utilities", "subsector": "Solar Power", "ai_tier": AITier.TIER2_INDIRECT},
    "221115": {"sector": "Utilities", "subsector": "Wind Power", "ai_tier": AITier.TIER2_INDIRECT},
    "221121": {"sector": "Utilities", "subsector": "Electric Bulk Power", "ai_tier": AITier.TIER2_INDIRECT},
    "221122": {"sector": "Utilities", "subsector": "Electric Power Distribution", "ai_tier": AITier.TIER3_MINIMAL},
    
    # Data Center Real Estate
    "531120": {"sector": "RealEstate", "subsector": "Lessors of Nonresidential (incl Data Centers)", "ai_tier": AITier.TIER2_INDIRECT},
    
    # =========================================================================
    # TIER 3: MINIMAL AI EXPOSURE
    # =========================================================================
    # Traditional Technology
    "334220": {"sector": "Technology", "subsector": "Radio/TV Broadcasting Equipment", "ai_tier": AITier.TIER3_MINIMAL},
    "334290": {"sector": "Technology", "subsector": "Other Communications Equipment", "ai_tier": AITier.TIER3_MINIMAL},
    "334310": {"sector": "Technology", "subsector": "Audio/Video Equipment", "ai_tier": AITier.TIER3_MINIMAL},
    
    # Financial Services
    "522110": {"sector": "Financials", "subsector": "Commercial Banking", "ai_tier": AITier.TIER3_MINIMAL},
    "522120": {"sector": "Financials", "subsector": "Savings Institutions", "ai_tier": AITier.TIER3_MINIMAL},
    "522130": {"sector": "Financials", "subsector": "Credit Unions", "ai_tier": AITier.TIER3_MINIMAL},
    "522210": {"sector": "Financials", "subsector": "Credit Card Issuing", "ai_tier": AITier.TIER3_MINIMAL},
    "522220": {"sector": "Financials", "subsector": "Sales Financing", "ai_tier": AITier.TIER3_MINIMAL},
    "522291": {"sector": "Financials", "subsector": "Consumer Lending", "ai_tier": AITier.TIER3_MINIMAL},
    "522292": {"sector": "Financials", "subsector": "Real Estate Credit", "ai_tier": AITier.TIER3_MINIMAL},
    "523110": {"sector": "Financials", "subsector": "Investment Banking", "ai_tier": AITier.TIER3_MINIMAL},
    "523120": {"sector": "Financials", "subsector": "Securities Brokerage", "ai_tier": AITier.TIER3_MINIMAL},
    "523130": {"sector": "Financials", "subsector": "Commodity Contracts", "ai_tier": AITier.TIER3_MINIMAL},
    "523910": {"sector": "Financials", "subsector": "Misc Intermediation", "ai_tier": AITier.TIER3_MINIMAL},
    "524113": {"sector": "Insurance", "subsector": "Direct Life Insurance", "ai_tier": AITier.TIER3_MINIMAL},
    "524114": {"sector": "Insurance", "subsector": "Direct Health Insurance", "ai_tier": AITier.TIER3_MINIMAL},
    "524126": {"sector": "Insurance", "subsector": "Direct P&C Insurance", "ai_tier": AITier.TIER3_MINIMAL},
    "524130": {"sector": "Insurance", "subsector": "Reinsurance", "ai_tier": AITier.TIER3_MINIMAL},
    
    # Energy
    "211120": {"sector": "Energy", "subsector": "Crude Petroleum Extraction", "ai_tier": AITier.TIER3_MINIMAL},
    "211130": {"sector": "Energy", "subsector": "Natural Gas Extraction", "ai_tier": AITier.TIER3_MINIMAL},
    "213111": {"sector": "Energy", "subsector": "Drilling Oil & Gas Wells", "ai_tier": AITier.TIER3_MINIMAL},
    "213112": {"sector": "Energy", "subsector": "Support Activities Oil & Gas", "ai_tier": AITier.TIER3_MINIMAL},
    "324110": {"sector": "Energy", "subsector": "Petroleum Refineries", "ai_tier": AITier.TIER3_MINIMAL},
    "486110": {"sector": "Energy", "subsector": "Pipeline Transportation Crude", "ai_tier": AITier.TIER3_MINIMAL},
    "486210": {"sector": "Energy", "subsector": "Pipeline Transportation Gas", "ai_tier": AITier.TIER3_MINIMAL},
    
    # Industrials
    "336411": {"sector": "Industrials", "subsector": "Aircraft Manufacturing", "ai_tier": AITier.TIER3_MINIMAL},
    "336412": {"sector": "Industrials", "subsector": "Aircraft Engine Manufacturing", "ai_tier": AITier.TIER3_MINIMAL},
    "336413": {"sector": "Industrials", "subsector": "Aircraft Parts", "ai_tier": AITier.TIER3_MINIMAL},
    "333120": {"sector": "Industrials", "subsector": "Construction Machinery", "ai_tier": AITier.TIER3_MINIMAL},
    "333131": {"sector": "Industrials", "subsector": "Mining Machinery", "ai_tier": AITier.TIER3_MINIMAL},
    "333914": {"sector": "Industrials", "subsector": "Measuring Instruments", "ai_tier": AITier.TIER3_MINIMAL},
    "336111": {"sector": "Industrials", "subsector": "Automobile Manufacturing", "ai_tier": AITier.TIER3_MINIMAL},
    "336112": {"sector": "Industrials", "subsector": "Light Truck Manufacturing", "ai_tier": AITier.TIER3_MINIMAL},
    "336120": {"sector": "Industrials", "subsector": "Heavy Truck Manufacturing", "ai_tier": AITier.TIER3_MINIMAL},
    "481111": {"sector": "Industrials", "subsector": "Scheduled Passenger Air", "ai_tier": AITier.TIER3_MINIMAL},
    "481112": {"sector": "Industrials", "subsector": "Scheduled Freight Air", "ai_tier": AITier.TIER3_MINIMAL},
    "482111": {"sector": "Industrials", "subsector": "Line-Haul Railroads", "ai_tier": AITier.TIER3_MINIMAL},
    "492110": {"sector": "Industrials", "subsector": "Couriers & Express", "ai_tier": AITier.TIER3_MINIMAL},
    
    # Materials
    "331110": {"sector": "Materials", "subsector": "Iron & Steel Mills", "ai_tier": AITier.TIER3_MINIMAL},
    "331313": {"sector": "Materials", "subsector": "Alumina Refining", "ai_tier": AITier.TIER3_MINIMAL},
    "331410": {"sector": "Materials", "subsector": "Nonferrous Metal Smelting", "ai_tier": AITier.TIER3_MINIMAL},
    "212210": {"sector": "Materials", "subsector": "Iron Ore Mining", "ai_tier": AITier.TIER3_MINIMAL},
    "212230": {"sector": "Materials", "subsector": "Copper Mining", "ai_tier": AITier.TIER3_MINIMAL},
    "212299": {"sector": "Materials", "subsector": "Other Metal Ore Mining", "ai_tier": AITier.TIER3_MINIMAL},
    "325110": {"sector": "Materials", "subsector": "Petrochemical Manufacturing", "ai_tier": AITier.TIER3_MINIMAL},
    "325120": {"sector": "Materials", "subsector": "Industrial Gas Manufacturing", "ai_tier": AITier.TIER3_MINIMAL},
    
    # Consumer Discretionary
    "441110": {"sector": "ConsumerDiscretionary", "subsector": "New Car Dealers", "ai_tier": AITier.TIER3_MINIMAL},
    "448110": {"sector": "ConsumerDiscretionary", "subsector": "Men's Clothing Stores", "ai_tier": AITier.TIER3_MINIMAL},
    "448120": {"sector": "ConsumerDiscretionary", "subsector": "Women's Clothing Stores", "ai_tier": AITier.TIER3_MINIMAL},
    "451110": {"sector": "ConsumerDiscretionary", "subsector": "Sporting Goods Stores", "ai_tier": AITier.TIER3_MINIMAL},
    "721110": {"sector": "ConsumerDiscretionary", "subsector": "Hotels", "ai_tier": AITier.TIER3_MINIMAL},
    "722511": {"sector": "ConsumerDiscretionary", "subsector": "Full-Service Restaurants", "ai_tier": AITier.TIER3_MINIMAL},
    
    # Real Estate (non-data center)
    "531110": {"sector": "RealEstate", "subsector": "Lessors of Residential", "ai_tier": AITier.TIER3_MINIMAL},
    "531130": {"sector": "RealEstate", "subsector": "Lessors of Miniwarehouses", "ai_tier": AITier.TIER3_MINIMAL},
    "531210": {"sector": "RealEstate", "subsector": "Real Estate Agents", "ai_tier": AITier.TIER3_MINIMAL},
    "531311": {"sector": "RealEstate", "subsector": "Residential Property Managers", "ai_tier": AITier.TIER3_MINIMAL},
    "531312": {"sector": "RealEstate", "subsector": "Nonresidential Property Managers", "ai_tier": AITier.TIER3_MINIMAL},
    
    # =========================================================================
    # DEFENSIVE SECTORS
    # =========================================================================
    # Healthcare - Pharma
    "325411": {"sector": "Healthcare", "subsector": "Medicinal & Botanical Manufacturing", "ai_tier": AITier.DEFENSIVE},
    "325412": {"sector": "Healthcare", "subsector": "Pharmaceutical Preparation", "ai_tier": AITier.DEFENSIVE},
    "325413": {"sector": "Healthcare", "subsector": "In-Vitro Diagnostic", "ai_tier": AITier.DEFENSIVE},
    "325414": {"sector": "Healthcare", "subsector": "Biological Products", "ai_tier": AITier.DEFENSIVE},
    
    # Healthcare - Services
    "621111": {"sector": "Healthcare", "subsector": "Offices of Physicians", "ai_tier": AITier.DEFENSIVE},
    "621210": {"sector": "Healthcare", "subsector": "Offices of Dentists", "ai_tier": AITier.DEFENSIVE},
    "621310": {"sector": "Healthcare", "subsector": "Offices of Chiropractors", "ai_tier": AITier.DEFENSIVE},
    "621410": {"sector": "Healthcare", "subsector": "Family Planning Centers", "ai_tier": AITier.DEFENSIVE},
    "621491": {"sector": "Healthcare", "subsector": "HMO Medical Centers", "ai_tier": AITier.DEFENSIVE},
    "621492": {"sector": "Healthcare", "subsector": "Kidney Dialysis Centers", "ai_tier": AITier.DEFENSIVE},
    "621610": {"sector": "Healthcare", "subsector": "Home Health Care Services", "ai_tier": AITier.DEFENSIVE},
    "622110": {"sector": "Healthcare", "subsector": "General Medical Hospitals", "ai_tier": AITier.DEFENSIVE},
    "622210": {"sector": "Healthcare", "subsector": "Psychiatric Hospitals", "ai_tier": AITier.DEFENSIVE},
    "622310": {"sector": "Healthcare", "subsector": "Specialty Hospitals", "ai_tier": AITier.DEFENSIVE},
    "623110": {"sector": "Healthcare", "subsector": "Nursing Care Facilities", "ai_tier": AITier.DEFENSIVE},
    
    # Consumer Staples - Food
    "311111": {"sector": "ConsumerStaples", "subsector": "Dog & Cat Food", "ai_tier": AITier.DEFENSIVE},
    "311211": {"sector": "ConsumerStaples", "subsector": "Flour Milling", "ai_tier": AITier.DEFENSIVE},
    "311212": {"sector": "ConsumerStaples", "subsector": "Rice Milling", "ai_tier": AITier.DEFENSIVE},
    "311221": {"sector": "ConsumerStaples", "subsector": "Wet Corn Milling", "ai_tier": AITier.DEFENSIVE},
    "311224": {"sector": "ConsumerStaples", "subsector": "Soybean Oil Mills", "ai_tier": AITier.DEFENSIVE},
    "311230": {"sector": "ConsumerStaples", "subsector": "Breakfast Cereal", "ai_tier": AITier.DEFENSIVE},
    "311410": {"sector": "ConsumerStaples", "subsector": "Frozen Food", "ai_tier": AITier.DEFENSIVE},
    "311421": {"sector": "ConsumerStaples", "subsector": "Fruit & Vegetable Canning", "ai_tier": AITier.DEFENSIVE},
    "311511": {"sector": "ConsumerStaples", "subsector": "Fluid Milk", "ai_tier": AITier.DEFENSIVE},
    "311520": {"sector": "ConsumerStaples", "subsector": "Ice Cream", "ai_tier": AITier.DEFENSIVE},
    "311612": {"sector": "ConsumerStaples", "subsector": "Meat Processed from Carcasses", "ai_tier": AITier.DEFENSIVE},
    "311710": {"sector": "ConsumerStaples", "subsector": "Seafood Preparation", "ai_tier": AITier.DEFENSIVE},
    "311812": {"sector": "ConsumerStaples", "subsector": "Commercial Bakeries", "ai_tier": AITier.DEFENSIVE},
    "311911": {"sector": "ConsumerStaples", "subsector": "Roasted Nuts & Peanut Butter", "ai_tier": AITier.DEFENSIVE},
    "311919": {"sector": "ConsumerStaples", "subsector": "Other Snack Food", "ai_tier": AITier.DEFENSIVE},
    
    # Consumer Staples - Beverages
    "312111": {"sector": "ConsumerStaples", "subsector": "Soft Drink Manufacturing", "ai_tier": AITier.DEFENSIVE},
    "312112": {"sector": "ConsumerStaples", "subsector": "Bottled Water", "ai_tier": AITier.DEFENSIVE},
    "312120": {"sector": "ConsumerStaples", "subsector": "Breweries", "ai_tier": AITier.DEFENSIVE},
    "312130": {"sector": "ConsumerStaples", "subsector": "Wineries", "ai_tier": AITier.DEFENSIVE},
    "312140": {"sector": "ConsumerStaples", "subsector": "Distilleries", "ai_tier": AITier.DEFENSIVE},
    
    # Consumer Staples - Household Products
    "325611": {"sector": "ConsumerStaples", "subsector": "Soap Manufacturing", "ai_tier": AITier.DEFENSIVE},
    "325612": {"sector": "ConsumerStaples", "subsector": "Polish Manufacturing", "ai_tier": AITier.DEFENSIVE},
    "325620": {"sector": "ConsumerStaples", "subsector": "Toilet Preparation", "ai_tier": AITier.DEFENSIVE},
    
    # Consumer Staples - Retail
    "445110": {"sector": "ConsumerStaples", "subsector": "Supermarkets", "ai_tier": AITier.DEFENSIVE},
    "445120": {"sector": "ConsumerStaples", "subsector": "Convenience Stores", "ai_tier": AITier.DEFENSIVE},
    "446110": {"sector": "ConsumerStaples", "subsector": "Pharmacies & Drug Stores", "ai_tier": AITier.DEFENSIVE},
    "446191": {"sector": "ConsumerStaples", "subsector": "Food Health Supplement Stores", "ai_tier": AITier.DEFENSIVE},
    
    # Consumer Staples - Tobacco
    "312230": {"sector": "ConsumerStaples", "subsector": "Tobacco Manufacturing", "ai_tier": AITier.DEFENSIVE},
}


# =============================================================================
# SECTION 4: COMPANY DATABASE
# =============================================================================

COMPANY_DATABASE = {
    # =========================================================================
    # TIER 1: DIRECT AI IMPACT - SEMICONDUCTORS
    # =========================================================================
    "NVDA": {
        "name": "NVIDIA Corporation", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "A", "beta": 2.15,
        "ai_revenue_pct": 85, "market_cap_bn": 2800,
        "description": "AI chip leader, data center GPUs"
    },
    "AMD": {
        "name": "Advanced Micro Devices", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "BBB", "beta": 1.95,
        "ai_revenue_pct": 45, "market_cap_bn": 220,
        "description": "AI challenger, MI300 accelerators"
    },
    "AVGO": {
        "name": "Broadcom Inc", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "BBB", "beta": 1.45,
        "ai_revenue_pct": 35, "market_cap_bn": 750,
        "description": "AI networking chips, custom ASICs"
    },
    "INTC": {
        "name": "Intel Corporation", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "A", "beta": 1.25,
        "ai_revenue_pct": 15, "market_cap_bn": 100,
        "description": "Legacy chipmaker, AI turnaround play"
    },
    "QCOM": {
        "name": "Qualcomm Inc", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "A", "beta": 1.55,
        "ai_revenue_pct": 25, "market_cap_bn": 180,
        "description": "Mobile AI chips, edge inference"
    },
    "MU": {
        "name": "Micron Technology", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "BBB", "beta": 1.75,
        "ai_revenue_pct": 40, "market_cap_bn": 110,
        "description": "HBM memory for AI accelerators"
    },
    "MRVL": {
        "name": "Marvell Technology", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "BBB", "beta": 1.85,
        "ai_revenue_pct": 50, "market_cap_bn": 65,
        "description": "Custom AI silicon, data center"
    },
    "ARM": {
        "name": "ARM Holdings", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 2.20,
        "ai_revenue_pct": 30, "market_cap_bn": 140,
        "description": "AI chip architecture licensing"
    },
    "TSM": {
        "name": "Taiwan Semiconductor", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "AA", "beta": 1.35,
        "ai_revenue_pct": 50, "market_cap_bn": 800,
        "description": "AI chip foundry, sole advanced node"
    },
    "ASML": {
        "name": "ASML Holding", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "AA", "beta": 1.40,
        "ai_revenue_pct": 60, "market_cap_bn": 350,
        "description": "EUV lithography monopoly for AI chips"
    },
    "LRCX": {
        "name": "Lam Research", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "A", "beta": 1.55,
        "ai_revenue_pct": 45, "market_cap_bn": 95,
        "description": "Semiconductor equipment, etch tools"
    },
    "AMAT": {
        "name": "Applied Materials", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "A", "beta": 1.50,
        "ai_revenue_pct": 40, "market_cap_bn": 150,
        "description": "Semiconductor equipment, deposition"
    },
    "KLAC": {
        "name": "KLA Corporation", "sector": "Semiconductors", "naic": "334413",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "A", "beta": 1.45,
        "ai_revenue_pct": 35, "market_cap_bn": 85,
        "description": "Semiconductor inspection equipment"
    },
    
    # =========================================================================
    # TIER 1: DIRECT AI IMPACT - AI SOFTWARE
    # =========================================================================
    "PLTR": {
        "name": "Palantir Technologies", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 2.50,
        "ai_revenue_pct": 90, "market_cap_bn": 150,
        "description": "AI/ML platform for enterprise"
    },
    "AI": {
        "name": "C3.ai Inc", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 2.80,
        "ai_revenue_pct": 95, "market_cap_bn": 4,
        "description": "Enterprise AI applications"
    },
    "PATH": {
        "name": "UiPath Inc", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 2.20,
        "ai_revenue_pct": 85, "market_cap_bn": 8,
        "description": "AI-powered automation"
    },
    "SNOW": {
        "name": "Snowflake Inc", "sector": "Software", "naic": "518210",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 2.10,
        "ai_revenue_pct": 60, "market_cap_bn": 55,
        "description": "AI data cloud platform"
    },
    "DDOG": {
        "name": "Datadog Inc", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 1.90,
        "ai_revenue_pct": 50, "market_cap_bn": 45,
        "description": "AI observability platform"
    },
    "MDB": {
        "name": "MongoDB Inc", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 2.05,
        "ai_revenue_pct": 45, "market_cap_bn": 25,
        "description": "AI application database"
    },
    "CRWD": {
        "name": "CrowdStrike Holdings", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER1_DIRECT, "rating": "NR", "beta": 1.75,
        "ai_revenue_pct": 70, "market_cap_bn": 75,
        "description": "AI cybersecurity platform"
    },
    
    # =========================================================================
    # TIER 2: INDIRECT AI IMPACT - HYPERSCALERS
    # =========================================================================
    "MSFT": {
        "name": "Microsoft Corporation", "sector": "Technology", "naic": "511210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "AAA", "beta": 1.12,
        "ai_revenue_pct": 25, "market_cap_bn": 3100,
        "description": "Azure AI, Copilot, OpenAI partner"
    },
    "GOOGL": {
        "name": "Alphabet Inc", "sector": "Technology", "naic": "518210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "AA", "beta": 1.18,
        "ai_revenue_pct": 20, "market_cap_bn": 2100,
        "description": "Google Cloud AI, Gemini"
    },
    "AMZN": {
        "name": "Amazon.com Inc", "sector": "Technology", "naic": "518210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "AA", "beta": 1.25,
        "ai_revenue_pct": 15, "market_cap_bn": 2000,
        "description": "AWS AI/ML services"
    },
    "META": {
        "name": "Meta Platforms", "sector": "Technology", "naic": "518210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "AA", "beta": 1.35,
        "ai_revenue_pct": 30, "market_cap_bn": 1400,
        "description": "Llama models, AI ad targeting"
    },
    "ORCL": {
        "name": "Oracle Corporation", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BBB", "beta": 1.05,
        "ai_revenue_pct": 15, "market_cap_bn": 380,
        "description": "Cloud AI infrastructure"
    },
    "CRM": {
        "name": "Salesforce Inc", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "A", "beta": 1.20,
        "ai_revenue_pct": 20, "market_cap_bn": 280,
        "description": "Einstein AI, Agentforce"
    },
    "NOW": {
        "name": "ServiceNow Inc", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "A", "beta": 1.25,
        "ai_revenue_pct": 25, "market_cap_bn": 200,
        "description": "AI workflow automation"
    },
    "ADBE": {
        "name": "Adobe Inc", "sector": "Software", "naic": "511210",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "A", "beta": 1.30,
        "ai_revenue_pct": 35, "market_cap_bn": 220,
        "description": "Generative AI creative tools"
    },
    
    # =========================================================================
    # TIER 2: INDIRECT AI IMPACT - DATA CENTER REITS
    # =========================================================================
    "EQIX": {
        "name": "Equinix Inc", "sector": "RealEstate", "naic": "531120",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BBB", "beta": 0.85,
        "ai_revenue_pct": 40, "market_cap_bn": 80,
        "description": "Data center REIT, AI workload hosting"
    },
    "DLR": {
        "name": "Digital Realty Trust", "sector": "RealEstate", "naic": "531120",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BBB", "beta": 0.80,
        "ai_revenue_pct": 35, "market_cap_bn": 45,
        "description": "Data center REIT"
    },
    
    # =========================================================================
    # TIER 2: INDIRECT AI IMPACT - UTILITIES (DATA CENTER POWER)
    # =========================================================================
    "VST": {
        "name": "Vistra Corp", "sector": "Utilities", "naic": "221112",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BB", "beta": 1.15,
        "ai_revenue_pct": 25, "market_cap_bn": 45,
        "description": "IPP, data center power deals"
    },
    "CEG": {
        "name": "Constellation Energy", "sector": "Utilities", "naic": "221113",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BBB", "beta": 0.95,
        "ai_revenue_pct": 20, "market_cap_bn": 75,
        "description": "Nuclear, Microsoft data center deal"
    },
    "NRG": {
        "name": "NRG Energy", "sector": "Utilities", "naic": "221112",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BB", "beta": 1.10,
        "ai_revenue_pct": 15, "market_cap_bn": 20,
        "description": "IPP, data center power exposure"
    },
    "AES": {
        "name": "AES Corporation", "sector": "Utilities", "naic": "221112",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BBB", "beta": 0.90,
        "ai_revenue_pct": 10, "market_cap_bn": 15,
        "description": "Renewable IPP, data center PPAs"
    },
    "OKLO": {
        "name": "Oklo Inc", "sector": "Utilities", "naic": "221113",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "NR", "beta": 2.50,
        "ai_revenue_pct": 100, "market_cap_bn": 3,
        "description": "SMR nuclear for data centers"
    },
    
    # =========================================================================
    # TIER 2: INDIRECT AI IMPACT - EV/AUTONOMY
    # =========================================================================
    "TSLA": {
        "name": "Tesla Inc", "sector": "ConsumerDiscretionary", "naic": "336111",
        "ai_tier": AITier.TIER2_INDIRECT, "rating": "BBB", "beta": 1.85,
        "ai_revenue_pct": 20, "market_cap_bn": 800,
        "description": "FSD, Dojo supercomputer"
    },
    
    # =========================================================================
    # TIER 3: MINIMAL AI EXPOSURE - TRADITIONAL TECH
    # =========================================================================
    "AAPL": {
        "name": "Apple Inc", "sector": "Technology", "naic": "334111",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "AA", "beta": 1.08,
        "ai_revenue_pct": 5, "market_cap_bn": 3500,
        "description": "Consumer devices, Apple Intelligence"
    },
    "CSCO": {
        "name": "Cisco Systems", "sector": "Technology", "naic": "334290",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "AA", "beta": 0.95,
        "ai_revenue_pct": 10, "market_cap_bn": 230,
        "description": "Networking equipment"
    },
    "IBM": {
        "name": "IBM Corporation", "sector": "Technology", "naic": "511210",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 0.85,
        "ai_revenue_pct": 15, "market_cap_bn": 200,
        "description": "Watson AI, hybrid cloud"
    },
    "HPQ": {
        "name": "HP Inc", "sector": "Technology", "naic": "334111",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "BBB", "beta": 1.05,
        "ai_revenue_pct": 5, "market_cap_bn": 35,
        "description": "PC/printers"
    },
    "DELL": {
        "name": "Dell Technologies", "sector": "Technology", "naic": "334111",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "BBB", "beta": 1.15,
        "ai_revenue_pct": 15, "market_cap_bn": 75,
        "description": "AI servers, storage"
    },
    
    # =========================================================================
    # TIER 3: MINIMAL AI EXPOSURE - FINANCIALS
    # =========================================================================
    "JPM": {
        "name": "JPMorgan Chase", "sector": "Financials", "naic": "522110",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.05,
        "ai_revenue_pct": 5, "market_cap_bn": 650,
        "description": "Largest US bank"
    },
    "BAC": {
        "name": "Bank of America", "sector": "Financials", "naic": "522110",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.10,
        "ai_revenue_pct": 3, "market_cap_bn": 350,
        "description": "Consumer banking"
    },
    "GS": {
        "name": "Goldman Sachs", "sector": "Financials", "naic": "523110",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.25,
        "ai_revenue_pct": 5, "market_cap_bn": 170,
        "description": "Investment banking"
    },
    "MS": {
        "name": "Morgan Stanley", "sector": "Financials", "naic": "523110",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.20,
        "ai_revenue_pct": 5, "market_cap_bn": 190,
        "description": "Wealth management, IB"
    },
    "C": {
        "name": "Citigroup", "sector": "Financials", "naic": "522110",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "BBB", "beta": 1.35,
        "ai_revenue_pct": 3, "market_cap_bn": 130,
        "description": "Global banking"
    },
    "WFC": {
        "name": "Wells Fargo", "sector": "Financials", "naic": "522110",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.15,
        "ai_revenue_pct": 3, "market_cap_bn": 230,
        "description": "Consumer banking"
    },
    "BLK": {
        "name": "BlackRock Inc", "sector": "Financials", "naic": "523920",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "AA", "beta": 1.15,
        "ai_revenue_pct": 5, "market_cap_bn": 140,
        "description": "Asset management"
    },
    
    # =========================================================================
    # TIER 3: MINIMAL AI EXPOSURE - ENERGY
    # =========================================================================
    "XOM": {
        "name": "Exxon Mobil", "sector": "Energy", "naic": "211120",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "AA", "beta": 0.85,
        "ai_revenue_pct": 1, "market_cap_bn": 520,
        "description": "Integrated oil & gas"
    },
    "CVX": {
        "name": "Chevron Corp", "sector": "Energy", "naic": "211120",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "AA", "beta": 0.90,
        "ai_revenue_pct": 1, "market_cap_bn": 280,
        "description": "Integrated oil & gas"
    },
    "COP": {
        "name": "ConocoPhillips", "sector": "Energy", "naic": "211120",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.10,
        "ai_revenue_pct": 1, "market_cap_bn": 130,
        "description": "E&P"
    },
    "OXY": {
        "name": "Occidental Petroleum", "sector": "Energy", "naic": "211120",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "BB", "beta": 1.45,
        "ai_revenue_pct": 1, "market_cap_bn": 50,
        "description": "E&P, Permian focus"
    },
    "SLB": {
        "name": "Schlumberger", "sector": "Energy", "naic": "213112",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.25,
        "ai_revenue_pct": 5, "market_cap_bn": 65,
        "description": "Oilfield services"
    },
    
    # =========================================================================
    # TIER 3: MINIMAL AI EXPOSURE - INDUSTRIALS
    # =========================================================================
    "BA": {
        "name": "Boeing Company", "sector": "Industrials", "naic": "336411",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "BBB", "beta": 1.35,
        "ai_revenue_pct": 2, "market_cap_bn": 120,
        "description": "Aerospace & defense"
    },
    "CAT": {
        "name": "Caterpillar Inc", "sector": "Industrials", "naic": "333120",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.00,
        "ai_revenue_pct": 3, "market_cap_bn": 180,
        "description": "Construction equipment"
    },
    "GE": {
        "name": "GE Aerospace", "sector": "Industrials", "naic": "336412",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "BBB", "beta": 1.15,
        "ai_revenue_pct": 5, "market_cap_bn": 190,
        "description": "Jet engines"
    },
    "HON": {
        "name": "Honeywell International", "sector": "Industrials", "naic": "334512",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 1.00,
        "ai_revenue_pct": 5, "market_cap_bn": 140,
        "description": "Industrial conglomerate"
    },
    "UPS": {
        "name": "United Parcel Service", "sector": "Industrials", "naic": "492110",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 0.95,
        "ai_revenue_pct": 3, "market_cap_bn": 115,
        "description": "Package delivery"
    },
    "RTX": {
        "name": "RTX Corporation", "sector": "Industrials", "naic": "336411",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 0.85,
        "ai_revenue_pct": 5, "market_cap_bn": 150,
        "description": "Defense & aerospace"
    },
    "LMT": {
        "name": "Lockheed Martin", "sector": "Industrials", "naic": "336411",
        "ai_tier": AITier.TIER3_MINIMAL, "rating": "A", "beta": 0.65,
        "ai_revenue_pct": 8, "market_cap_bn": 130,
        "description": "Defense contractor"
    },
    
    # =========================================================================
    # DEFENSIVE - HEALTHCARE
    # =========================================================================
    "JNJ": {
        "name": "Johnson & Johnson", "sector": "Healthcare", "naic": "325412",
        "ai_tier": AITier.DEFENSIVE, "rating": "AAA", "beta": 0.55,
        "ai_revenue_pct": 2, "market_cap_bn": 380,
        "description": "Pharma, medtech"
    },
    "UNH": {
        "name": "UnitedHealth Group", "sector": "Healthcare", "naic": "524114",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.65,
        "ai_revenue_pct": 5, "market_cap_bn": 480,
        "description": "Health insurance"
    },
    "PFE": {
        "name": "Pfizer Inc", "sector": "Healthcare", "naic": "325412",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.60,
        "ai_revenue_pct": 3, "market_cap_bn": 150,
        "description": "Pharmaceuticals"
    },
    "MRK": {
        "name": "Merck & Co", "sector": "Healthcare", "naic": "325412",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.55,
        "ai_revenue_pct": 5, "market_cap_bn": 280,
        "description": "Pharmaceuticals"
    },
    "ABBV": {
        "name": "AbbVie Inc", "sector": "Healthcare", "naic": "325412",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.60,
        "ai_revenue_pct": 3, "market_cap_bn": 310,
        "description": "Pharmaceuticals"
    },
    "LLY": {
        "name": "Eli Lilly", "sector": "Healthcare", "naic": "325412",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.50,
        "ai_revenue_pct": 10, "market_cap_bn": 750,
        "description": "Pharma, AI drug discovery"
    },
    "TMO": {
        "name": "Thermo Fisher Scientific", "sector": "Healthcare", "naic": "334516",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.80,
        "ai_revenue_pct": 5, "market_cap_bn": 200,
        "description": "Life sciences equipment"
    },
    "BMY": {
        "name": "Bristol-Myers Squibb", "sector": "Healthcare", "naic": "325412",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.50,
        "ai_revenue_pct": 3, "market_cap_bn": 100,
        "description": "Pharmaceuticals"
    },
    
    # =========================================================================
    # DEFENSIVE - CONSUMER STAPLES
    # =========================================================================
    "PG": {
        "name": "Procter & Gamble", "sector": "ConsumerStaples", "naic": "325611",
        "ai_tier": AITier.DEFENSIVE, "rating": "AA", "beta": 0.45,
        "ai_revenue_pct": 1, "market_cap_bn": 390,
        "description": "Consumer products"
    },
    "KO": {
        "name": "Coca-Cola Co", "sector": "ConsumerStaples", "naic": "312111",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.55,
        "ai_revenue_pct": 1, "market_cap_bn": 260,
        "description": "Beverages"
    },
    "PEP": {
        "name": "PepsiCo Inc", "sector": "ConsumerStaples", "naic": "312111",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.55,
        "ai_revenue_pct": 2, "market_cap_bn": 210,
        "description": "Beverages & snacks"
    },
    "WMT": {
        "name": "Walmart Inc", "sector": "ConsumerStaples", "naic": "445110",
        "ai_tier": AITier.DEFENSIVE, "rating": "AA", "beta": 0.50,
        "ai_revenue_pct": 3, "market_cap_bn": 680,
        "description": "Retail"
    },
    "COST": {
        "name": "Costco Wholesale", "sector": "ConsumerStaples", "naic": "445110",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.65,
        "ai_revenue_pct": 2, "market_cap_bn": 380,
        "description": "Warehouse retail"
    },
    "PM": {
        "name": "Philip Morris International", "sector": "ConsumerStaples", "naic": "312230",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.55,
        "ai_revenue_pct": 1, "market_cap_bn": 190,
        "description": "Tobacco"
    },
    "MO": {
        "name": "Altria Group", "sector": "ConsumerStaples", "naic": "312230",
        "ai_tier": AITier.DEFENSIVE, "rating": "BBB", "beta": 0.55,
        "ai_revenue_pct": 1, "market_cap_bn": 90,
        "description": "Tobacco"
    },
    "CL": {
        "name": "Colgate-Palmolive", "sector": "ConsumerStaples", "naic": "325611",
        "ai_tier": AITier.DEFENSIVE, "rating": "AA", "beta": 0.50,
        "ai_revenue_pct": 1, "market_cap_bn": 75,
        "description": "Consumer products"
    },
    "KMB": {
        "name": "Kimberly-Clark", "sector": "ConsumerStaples", "naic": "322121",
        "ai_tier": AITier.DEFENSIVE, "rating": "A", "beta": 0.45,
        "ai_revenue_pct": 1, "market_cap_bn": 45,
        "description": "Consumer products"
    },
}


# =============================================================================
# SECTION 5: SHOCK CONFIGURATION TABLES
# =============================================================================

# -----------------------------------------------------------------------------
# EQUITY SHOCKS
# -----------------------------------------------------------------------------

EQUITY_INDEX_SHOCKS = {
    "SPX": {"mod": -22, "sev": -33, "description": "S&P 500 anchor"},
    "NDX": {"mod": -30, "sev": -43, "description": "Nasdaq-100 (1.36x beta)"},
    "QQQ": {"mod": -30, "sev": -43, "description": "Nasdaq-100 ETF"},
    "RTY": {"mod": -18, "sev": -27, "description": "Russell 2000 (0.82x beta)"},
    "DJI": {"mod": -18, "sev": -28, "description": "Dow Jones (0.85x beta)"},
    "EAFE": {"mod": -14, "sev": -21, "description": "Developed ex-US (0.65x beta)"},
    "EEM": {"mod": -19, "sev": -28, "description": "EM equities (0.85x beta)"},
    "VGK": {"mod": -16, "sev": -24, "description": "Europe (0.72x beta)"},
    "EWJ": {"mod": -12, "sev": -18, "description": "Japan (0.55x beta)"},
    "FXI": {"mod": -15, "sev": -22, "description": "China (0.68x beta)"},
}

EQUITY_SECTOR_SHOCKS = {
    # AI-impacted sectors
    "Semiconductors": {"beta": 1.90, "mod": -42, "sev": -63, "ai_overlay_mod": -8, "ai_overlay_sev": -12},
    "Software": {"beta": 1.35, "mod": -30, "sev": -45, "ai_overlay_mod": -3, "ai_overlay_sev": -5},
    "Technology": {"beta": 1.45, "mod": -32, "sev": -48, "ai_overlay_mod": -5, "ai_overlay_sev": -8},
    
    # Market sectors
    "CommunicationServices": {"beta": 1.25, "mod": -28, "sev": -41, "ai_overlay_mod": 0, "ai_overlay_sev": 0},
    "ConsumerDiscretionary": {"beta": 1.15, "mod": -25, "sev": -38, "ai_overlay_mod": 0, "ai_overlay_sev": 0},
    "Financials": {"beta": 0.95, "mod": -21, "sev": -31, "ai_overlay_mod": 0, "ai_overlay_sev": 0},
    "Industrials": {"beta": 0.90, "mod": -20, "sev": -30, "ai_overlay_mod": 0, "ai_overlay_sev": 0},
    "Materials": {"beta": 0.85, "mod": -19, "sev": -28, "ai_overlay_mod": 0, "ai_overlay_sev": 0},
    "RealEstate": {"beta": 0.80, "mod": -18, "sev": -26, "ai_overlay_mod": 0, "ai_overlay_sev": 0},
    "Energy": {"beta": 0.75, "mod": -17, "sev": -25, "ai_overlay_mod": 0, "ai_overlay_sev": 0},
    
    # Defensive sectors
    "Healthcare": {"beta": 0.65, "mod": -14, "sev": -21, "ai_overlay_mod": 3, "ai_overlay_sev": 5},
    "Utilities": {"beta": 0.50, "mod": -11, "sev": -17, "ai_overlay_mod": 2, "ai_overlay_sev": 4},
    "ConsumerStaples": {"beta": 0.45, "mod": -10, "sev": -15, "ai_overlay_mod": 4, "ai_overlay_sev": 6},
}

EQUITY_BINDING_CONSTRAINTS = {
    # Tier 1 AI - Semiconductors (epicenter)
    "NVDA": {"mod": -40, "sev": -50, "rationale": "AI chip leader, highest AI revenue exposure"},
    "AMD": {"mod": -40, "sev": -50, "rationale": "AI challenger, MI300 competition"},
    "AVGO": {"mod": -40, "sev": -50, "rationale": "AI networking, custom ASIC exposure"},
    "MRVL": {"mod": -38, "sev": -48, "rationale": "Custom AI silicon"},
    "ARM": {"mod": -42, "sev": -52, "rationale": "AI architecture, premium valuation"},
    "TSM": {"mod": -35, "sev": -45, "rationale": "AI foundry, sole advanced node"},
    "ASML": {"mod": -32, "sev": -42, "rationale": "AI capex exposure via EUV"},
    "MU": {"mod": -35, "sev": -45, "rationale": "HBM memory for AI"},
    "LRCX": {"mod": -30, "sev": -40, "rationale": "AI capex exposure"},
    "AMAT": {"mod": -30, "sev": -40, "rationale": "AI capex exposure"},
    
    # Tier 1 AI - Software
    "PLTR": {"mod": -45, "sev": -55, "rationale": "Pure-play AI platform, premium valuation"},
    "SNOW": {"mod": -42, "sev": -52, "rationale": "AI data cloud premium"},
    "AI": {"mod": -50, "sev": -60, "rationale": "Pure AI play, speculative"},
    "PATH": {"mod": -40, "sev": -50, "rationale": "AI automation"},
    
    # Tier 2 AI - Hyperscalers
    "MSFT": {"mod": -35, "sev": -45, "rationale": "Azure AI, Copilot, OpenAI"},
    "GOOGL": {"mod": -35, "sev": -45, "rationale": "Google AI, Gemini"},
    "AMZN": {"mod": -35, "sev": -45, "rationale": "AWS AI"},
    "META": {"mod": -38, "sev": -48, "rationale": "Llama, AI capex"},
    
    # Tier 2 AI - Utilities (data center power)
    "VST": {"mod": -15, "sev": -30, "rationale": "Data center power deals"},
    "CEG": {"mod": -15, "sev": -30, "rationale": "Nuclear for AI data centers"},
    "OKLO": {"mod": -45, "sev": -55, "rationale": "SMR for AI, speculative"},
    
    # Tier 2 AI - Data center REITs
    "EQIX": {"mod": -18, "sev": -28, "rationale": "Data center REIT, AI workloads"},
    "DLR": {"mod": -17, "sev": -27, "rationale": "Data center REIT"},
    
    # Tier 2 AI - Autonomy
    "TSLA": {"mod": -32, "sev": -42, "rationale": "FSD, Dojo, AI premium"},
    
    # Tier 3 - Traditional tech (less impacted)
    "AAPL": {"mod": -28, "sev": -38, "rationale": "Limited AI exposure, consumer focus"},
}


# -----------------------------------------------------------------------------
# CREDIT SHOCKS
# -----------------------------------------------------------------------------

# Corporate: AAA to D/NR → 40-150bp (mod) / 75-500bp (sev)
CORP_RATING_SHOCKS = {
    "AAA": {"mod": 40, "sev": 75, "derivation": "Historical: COVID=35bp, GFC=80bp, blend=40/75"},
    "AA+": {"mod": 45, "sev": 85, "derivation": "AAA + 5/10bp notch"},
    "AA": {"mod": 50, "sev": 95, "derivation": "Historical: COVID=45bp, GFC=95bp"},
    "AA-": {"mod": 55, "sev": 105, "derivation": "AA + 5/10bp notch"},
    "A+": {"mod": 62, "sev": 115, "derivation": "Interpolated AA- to A"},
    "A": {"mod": 70, "sev": 130, "derivation": "Historical: COVID=60bp, GFC=140bp, blend=70/130"},
    "A-": {"mod": 78, "sev": 145, "derivation": "A + 8/15bp notch"},
    "BBB+": {"mod": 88, "sev": 165, "derivation": "Interpolated A- to BBB"},
    "BBB": {"mod": 100, "sev": 190, "derivation": "Historical: COVID=90bp, GFC=200bp, benchmark"},
    "BBB-": {"mod": 115, "sev": 220, "derivation": "BBB + 15/30bp crossover risk premium"},
    "BB+": {"mod": 135, "sev": 260, "derivation": "IG→HY transition, 20bp jump"},
    "BB": {"mod": 155, "sev": 310, "derivation": "Historical: COVID=140bp, GFC=350bp"},
    "BB-": {"mod": 175, "sev": 365, "derivation": "BB + 20/55bp notch"},
    "B+": {"mod": 195, "sev": 420, "derivation": "Approaching distress territory"},
    "B": {"mod": 220, "sev": 480, "derivation": "Historical: COVID=200bp, GFC=500bp"},
    "B-": {"mod": 250, "sev": 540, "derivation": "B + 30/60bp elevated default risk"},
    "CCC+": {"mod": 285, "sev": 600, "derivation": "Distress begins"},
    "CCC": {"mod": 320, "sev": 670, "derivation": "Historical: COVID=300bp, GFC=700bp"},
    "CCC-": {"mod": 360, "sev": 750, "derivation": "Near-default"},
    "CC": {"mod": 410, "sev": 850, "derivation": "Highly speculative"},
    "C": {"mod": 470, "sev": 960, "derivation": "Imminent default"},
    "D": {"mod": 550, "sev": 1100, "derivation": "Defaulted, recovery value"},
    "NR": {"mod": 150, "sev": 500, "derivation": "Assumed BB- equivalent"},
}

# CDS: 5-50bp (mod) / 30-300bp (sev)
CDS_RATING_SHOCKS = {
    "AAA": {"mod": 5, "sev": 30, "derivation": "Minimal default risk, liquidity premium"},
    "AA+": {"mod": 7, "sev": 38, "derivation": "AAA + 2/8bp"},
    "AA": {"mod": 9, "sev": 48, "derivation": "Historical: COVID=8bp, GFC=50bp"},
    "AA-": {"mod": 11, "sev": 58, "derivation": "AA + 2/10bp"},
    "A+": {"mod": 14, "sev": 72, "derivation": "Interpolated"},
    "A": {"mod": 18, "sev": 90, "derivation": "Historical: COVID=15bp, GFC=100bp"},
    "A-": {"mod": 22, "sev": 108, "derivation": "A + 4/18bp"},
    "BBB+": {"mod": 28, "sev": 135, "derivation": "Interpolated"},
    "BBB": {"mod": 35, "sev": 168, "derivation": "Historical: COVID=30bp, GFC=180bp"},
    "BBB-": {"mod": 42, "sev": 205, "derivation": "BBB + 7/37bp crossover"},
    "BB+": {"mod": 55, "sev": 260, "derivation": "Jump to HY"},
    "BB": {"mod": 72, "sev": 330, "derivation": "Historical: COVID=65bp, GFC=350bp"},
    "BB-": {"mod": 92, "sev": 410, "derivation": "BB + 20/80bp"},
    "B+": {"mod": 115, "sev": 500, "derivation": "Elevated default"},
    "B": {"mod": 145, "sev": 600, "derivation": "Historical: COVID=130bp, GFC=650bp"},
    "B-": {"mod": 180, "sev": 720, "derivation": "Near-distress"},
    "CCC+": {"mod": 225, "sev": 860, "derivation": "Distressed"},
    "CCC": {"mod": 280, "sev": 1020, "derivation": "Historical: COVID=250bp, GFC=1100bp"},
    "CCC-": {"mod": 350, "sev": 1200, "derivation": "Near-default"},
    "CC": {"mod": 440, "sev": 1450, "derivation": "Highly distressed"},
    "C": {"mod": 550, "sev": 1750, "derivation": "Imminent default"},
    "D": {"mod": 700, "sev": 2200, "derivation": "Default, recovery CDS"},
    "NR": {"mod": 50, "sev": 300, "derivation": "Assumed BB- equivalent"},
}

# Index: IG 15/125, HY 50/350
INDEX_SHOCKS = {
    "CDX.NA.IG": {"mod": 15, "sev": 125, "derivation": "Historical: COVID=120bp, 2022=80bp, blend"},
    "CDX.NA.HY": {"mod": 50, "sev": 350, "derivation": "Historical: COVID=600bp, 2022=280bp, blend"},
    "CDX.NA.IG.HVOL": {"mod": 25, "sev": 200, "derivation": "High vol names, 1.6x IG"},
    "CDX.EM": {"mod": 35, "sev": 280, "derivation": "EM sovereign/corporate blend"},
    "iTraxx.Europe": {"mod": 15, "sev": 120, "derivation": "Similar to CDX IG"},
    "iTraxx.Crossover": {"mod": 45, "sev": 320, "derivation": "BB/BBB names"},
    "iTraxx.Asia": {"mod": 20, "sev": 160, "derivation": "Asia HG focus"},
    "LCDX": {"mod": 55, "sev": 380, "derivation": "Loan CDS, secured recovery"},
}

# Tenor multipliers: 3M to 30Y
TENOR_MULTIPLIERS = {
    "3M": {"mult": 0.30, "derivation": "Minimal duration, liquidity premium only"},
    "6M": {"mult": 0.40, "derivation": "Short-end"},
    "1Y": {"mult": 0.55, "derivation": "1Y point"},
    "2Y": {"mult": 0.75, "derivation": "Below belly"},
    "3Y": {"mult": 0.88, "derivation": "Approaching belly"},
    "4Y": {"mult": 0.95, "derivation": "Near belly"},
    "5Y": {"mult": 1.00, "derivation": "Benchmark, most liquid"},
    "7Y": {"mult": 1.05, "derivation": "Peak spread sensitivity"},
    "10Y": {"mult": 1.02, "derivation": "Long-end benchmark"},
    "15Y": {"mult": 0.95, "derivation": "Duration rolldown begins"},
    "20Y": {"mult": 0.90, "derivation": "Ultra-long, less liquid"},
    "25Y": {"mult": 0.87, "derivation": "Ultra-long"},
    "30Y": {"mult": 0.85, "derivation": "Max tenor, duration compression"},
}

# Sector overlays: Tech +25/+50
CREDIT_SECTOR_OVERLAYS = {
    "Semiconductors": {"mod": 25, "sev": 50, "derivation": "AI epicenter"},
    "Technology": {"mod": 25, "sev": 50, "derivation": "Tech sector overlay"},
    "Software": {"mod": 20, "sev": 40, "derivation": "AI applications"},
    "Healthcare": {"mod": -10, "sev": -20, "derivation": "Defensive"},
    "ConsumerStaples": {"mod": -12, "sev": -25, "derivation": "Defensive"},
    "Utilities": {"mod": -5, "sev": -10, "derivation": "Regulated, defensive"},
    "Financials": {"mod": 5, "sev": 15, "derivation": "Credit exposure"},
    "Energy": {"mod": 8, "sev": 20, "derivation": "Commodity correlation"},
    "Industrials": {"mod": 5, "sev": 12, "derivation": "Cyclical"},
    "ConsumerDiscretionary": {"mod": 10, "sev": 25, "derivation": "Cyclical"},
    "RealEstate": {"mod": 12, "sev": 30, "derivation": "Rate sensitive"},
    "CommunicationServices": {"mod": 8, "sev": 18, "derivation": "Mixed"},
    "Materials": {"mod": 6, "sev": 15, "derivation": "Cyclical"},
    "Telecom": {"mod": 3, "sev": 8, "derivation": "Defensive infrastructure"},
    "Insurance": {"mod": 5, "sev": 12, "derivation": "Credit exposure"},
}

# AI Tier overlays for credit
CREDIT_AI_TIER_OVERLAYS = {
    AITier.TIER1_DIRECT: {"mod": 35, "sev": 75, "derivation": "Direct AI revenue exposure"},
    AITier.TIER2_INDIRECT: {"mod": 20, "sev": 45, "derivation": "Indirect AI exposure"},
    AITier.TIER3_MINIMAL: {"mod": 0, "sev": 0, "derivation": "No AI overlay"},
    AITier.DEFENSIVE: {"mod": -8, "sev": -18, "derivation": "Flight to quality"},
}


# -----------------------------------------------------------------------------
# COMMODITY SHOCKS
# -----------------------------------------------------------------------------

COMMODITY_SHOCKS = {
    # Energy - demand destruction from growth concerns
    "WTI": {"mod": -12, "sev": -25, "type": "Energy", "derivation": "Growth demand destruction"},
    "BRENT": {"mod": -11, "sev": -23, "type": "Energy", "derivation": "WTI * 0.92 spread"},
    "NG": {"mod": -8, "sev": -18, "type": "Energy", "derivation": "Seasonal, data center offset"},
    "RBOB": {"mod": -14, "sev": -28, "type": "Energy", "derivation": "Gasoline, demand sensitive"},
    "HO": {"mod": -13, "sev": -26, "type": "Energy", "derivation": "Heating oil"},
    "GASOIL": {"mod": -11, "sev": -24, "type": "Energy", "derivation": "European distillate"},
    
    # Precious - safe haven bid
    "XAU": {"mod": 5, "sev": 12, "type": "Precious", "derivation": "Safe haven, rate cut expectations"},
    "XAG": {"mod": 2, "sev": 6, "type": "Precious", "derivation": "Industrial/precious mix"},
    "XPT": {"mod": -3, "sev": -8, "type": "Precious", "derivation": "Auto demand weakness"},
    "XPD": {"mod": -5, "sev": -12, "type": "Precious", "derivation": "Auto catalyst, industrial"},
    
    # Industrial - growth concerns, AI capex pause
    "HG": {"mod": -18, "sev": -32, "type": "Industrial", "derivation": "Growth proxy, construction"},
    "ALI": {"mod": -15, "sev": -28, "type": "Industrial", "derivation": "China construction"},
    "NI": {"mod": -12, "sev": -25, "type": "Industrial", "derivation": "Stainless steel demand"},
    "ZN": {"mod": -10, "sev": -22, "type": "Industrial", "derivation": "Galvanized steel"},
    "PB": {"mod": -8, "sev": -18, "type": "Industrial", "derivation": "Battery demand"},
    "IRON": {"mod": -20, "sev": -35, "type": "Industrial", "derivation": "China steel, construction"},
    
    # Agricultural - less correlated
    "CORN": {"mod": -5, "sev": -12, "type": "Agricultural", "derivation": "Weather dependent"},
    "WHEAT": {"mod": -4, "sev": -10, "type": "Agricultural", "derivation": "Weather dependent"},
    "SOYBEAN": {"mod": -6, "sev": -14, "type": "Agricultural", "derivation": "Weather/trade"},
}


# -----------------------------------------------------------------------------
# RATES SHOCKS
# -----------------------------------------------------------------------------

RATES_SHOCKS = {
    # Treasuries - flight to quality
    "UST.2Y": {"mod": -25, "sev": -60, "derivation": "Front-end, Fed cut pricing"},
    "UST.5Y": {"mod": -35, "sev": -80, "derivation": "Belly rally"},
    "UST.10Y": {"mod": -40, "sev": -90, "derivation": "Benchmark, flight to quality"},
    "UST.30Y": {"mod": -30, "sev": -70, "derivation": "Long-end, duration demand"},
    
    # Curve movements
    "UST.2s10s": {"mod": 15, "sev": 30, "derivation": "Bull steepening (cuts priced)"},
    "UST.5s30s": {"mod": -5, "sev": -10, "derivation": "Long-end flattening"},
    
    # SOFR
    "SOFR.3M": {"mod": -15, "sev": -35, "derivation": "Front-end easing"},
    "SOFR.1Y": {"mod": -25, "sev": -55, "derivation": "Rate cut expectations"},
    "SOFR.5Y": {"mod": -35, "sev": -75, "derivation": "Terminal rate repricing"},
    
    # International (spread to UST)
    "BUND.10Y": {"mod": -30, "sev": -70, "derivation": "German flight to quality"},
    "GILT.10Y": {"mod": -25, "sev": -60, "derivation": "UK safe haven"},
    "JGB.10Y": {"mod": -5, "sev": -15, "derivation": "BOJ YCC, limited move"},
}


# -----------------------------------------------------------------------------
# FX SHOCKS
# -----------------------------------------------------------------------------

FX_SHOCKS = {
    # DXY anchor
    "DXY": {"mod": 3.0, "sev": 6.0, "derivation": "USD safe haven bid"},
    
    # G10 vs USD
    "EURUSD": {"mod": -2.9, "sev": -5.7, "derivation": "DXY * -0.95 (EUR weight)"},
    "USDJPY": {"mod": -1.8, "sev": -3.6, "derivation": "JPY safe haven partially offsets USD"},
    "GBPUSD": {"mod": -3.2, "sev": -6.3, "derivation": "GBP beta 1.05"},
    "USDCHF": {"mod": 2.1, "sev": 4.2, "derivation": "CHF safe haven, partial offset"},
    "AUDUSD": {"mod": -3.9, "sev": -7.8, "derivation": "Risk-off, commodity currency"},
    "NZDUSD": {"mod": -4.1, "sev": -8.1, "derivation": "AUD proxy, smaller"},
    "USDCAD": {"mod": 3.3, "sev": 6.6, "derivation": "Oil correlation"},
    "USDSEK": {"mod": 4.2, "sev": 8.4, "derivation": "Risk-off"},
    "USDNOK": {"mod": 4.5, "sev": 9.0, "derivation": "Oil + risk-off"},
    
    # EM vs USD (higher beta)
    "USDMXN": {"mod": 4.5, "sev": 9.0, "derivation": "EM beta 1.50"},
    "USDBRL": {"mod": 5.4, "sev": 10.8, "derivation": "EM beta 1.80"},
    "USDZAR": {"mod": 5.0, "sev": 9.9, "derivation": "EM beta 1.65"},
    "USDTRY": {"mod": 4.2, "sev": 8.4, "derivation": "EM beta 1.40"},
    "USDCNH": {"mod": 2.4, "sev": 4.8, "derivation": "CNH beta 0.80, managed"},
    "USDINR": {"mod": 1.8, "sev": 3.6, "derivation": "INR beta 0.60, RBI intervention"},
    "USDKRW": {"mod": 3.6, "sev": 7.2, "derivation": "KRW beta 1.20, tech exposure"},
    "USDTWD": {"mod": 3.0, "sev": 6.0, "derivation": "TWD beta 1.00, TSMC exposure"},
}


# -----------------------------------------------------------------------------
# VOLATILITY SHOCKS
# -----------------------------------------------------------------------------

VOL_SHOCKS = {
    # VIX anchor
    "VIX": {"mod_level": 38, "sev_level": 58, "mod_chg": 19.2, "sev_chg": 39.2,
            "derivation": "Historical: COVID=82, 2022=36, target=38/58"},
    "VVIX": {"mod_level": 110, "sev_level": 135, "derivation": "Vol-of-vol expansion"},
    
    # Term structure
    "VIX.1M": {"mod_level": 42, "sev_level": 65, "derivation": "Front-end inversion"},
    "VIX.3M": {"mod_level": 38, "sev_level": 58, "derivation": "Benchmark"},
    "VIX.6M": {"mod_level": 35, "sev_level": 52, "derivation": "Backwardation"},
    
    # Skew
    "SPX.25P.3M": {"mod": 8, "sev": 18, "derivation": "Put skew steepens (vol points)"},
    "SPX.25C.3M": {"mod": -2, "sev": -5, "derivation": "Call skew flattens"},
    "SPX.RR.25.3M": {"mod": -10, "sev": -23, "derivation": "Risk reversal widens"},
    
    # Correlation
    "SPX.CORR.IMPLIED": {"mod_level": 0.65, "sev_level": 0.85, 
                         "derivation": "Historical: COVID=0.80, 2022=0.60"},
    
    # Single stock vol (AI names)
    "NVDA.ATM.1M": {"mod_mult": 1.8, "sev_mult": 2.5, "derivation": "AI epicenter vol expansion"},
    "AMD.ATM.1M": {"mod_mult": 1.7, "sev_mult": 2.3, "derivation": "AI challenger"},
    "TSLA.ATM.1M": {"mod_mult": 1.5, "sev_mult": 2.0, "derivation": "High beta + autonomy"},
}


# =============================================================================
# SECTION 6: DATA CLASSES
# =============================================================================

@dataclass
class Driver:
    """Represents any stress scenario driver across asset classes."""
    driver_id: str
    name: str
    asset_class: AssetClass
    
    # Shock values
    shock_mod: float = 0.0
    shock_sev: float = 0.0
    
    # Current level (for context)
    current_level: float = 0.0
    stressed_level_mod: float = 0.0
    stressed_level_sev: float = 0.0
    
    # Classification
    sector: Optional[str] = None
    subsector: Optional[str] = None
    ai_tier: AITier = AITier.TIER3_MINIMAL
    naic_code: Optional[str] = None
    
    # Credit-specific
    credit_type: Optional[CreditType] = None
    rating: Optional[str] = None
    tenor: Optional[str] = None
    
    # Equity-specific
    beta: float = 1.0
    ticker: Optional[str] = None
    
    # Commodity-specific
    commodity_type: Optional[str] = None
    
    # Derivation
    parent_driver: Optional[str] = None
    base_shock_mod: float = 0.0
    base_shock_sev: float = 0.0
    overlay_mod: float = 0.0
    overlay_sev: float = 0.0
    multiplier: float = 1.0
    derivation: str = ""
    
    # Metadata
    is_anchor: bool = False
    is_binding_constraint: bool = False
    
    def compute_derivation_string(self):
        """Build human-readable derivation trail."""
        parts = []
        if self.is_anchor:
            parts.append("ANCHOR")
        elif self.is_binding_constraint:
            parts.append("BINDING_CONSTRAINT")
        else:
            parts.append(f"Base({self.base_shock_mod:.1f}/{self.base_shock_sev:.1f})")
            if self.multiplier != 1.0:
                parts.append(f"× Mult({self.multiplier:.2f})")
            if self.overlay_mod != 0 or self.overlay_sev != 0:
                parts.append(f"+ Overlay({self.overlay_mod:+.1f}/{self.overlay_sev:+.1f})")
        parts.append(f"= {self.shock_mod:.1f}/{self.shock_sev:.1f}")
        self.derivation = " ".join(parts)


@dataclass
class ValidationResult:
    """Result from a single validation check."""
    check_name: str
    driver_id: str
    severity: ValidationSeverity
    message: str
    expected: Optional[Any] = None
    actual: Optional[Any] = None


# =============================================================================
# SECTION 7: UNIFIED STRESS ENGINE
# =============================================================================

class UnifiedStressEngine:
    """
    Main engine for building and validating stress scenarios.
    Handles all asset classes with consistent methodology.
    """
    
    def __init__(self):
        self.drivers: Dict[str, Driver] = {}
        self.validation_results: List[ValidationResult] = []
        self.scenario_name = "AI Bubble Correction"
        self.generated_at = datetime.now()
    
    # -------------------------------------------------------------------------
    # BUILD METHODS
    # -------------------------------------------------------------------------
    
    def build_full_scenario(self):
        """Build complete scenario across all asset classes."""
        print("Building AI Bubble Correction Stress Scenario...")
        print("=" * 70)
        
        # 1. Equity
        print("\n1. Building EQUITY drivers...")
        self._build_equity_drivers()
        
        # 2. Credit
        print("2. Building CREDIT drivers...")
        self._build_credit_drivers()
        
        # 3. Commodities
        print("3. Building COMMODITY drivers...")
        self._build_commodity_drivers()
        
        # 4. Rates
        print("4. Building RATES drivers...")
        self._build_rates_drivers()
        
        # 5. FX
        print("5. Building FX drivers...")
        self._build_fx_drivers()
        
        # 6. Volatility
        print("6. Building VOLATILITY drivers...")
        self._build_vol_drivers()
        
        # 7. Validate
        print("\n7. Running validation...")
        self._run_validation()
        
        print(f"\n{'=' * 70}")
        print(f"SCENARIO BUILD COMPLETE")
        print(f"Total drivers: {len(self.drivers)}")
        print(f"Validation errors: {sum(1 for v in self.validation_results if v.severity == ValidationSeverity.ERROR)}")
        print(f"Validation warnings: {sum(1 for v in self.validation_results if v.severity == ValidationSeverity.WARNING)}")
    
    def _build_equity_drivers(self):
        """Build all equity drivers."""
        # Indices
        for idx, shock in EQUITY_INDEX_SHOCKS.items():
            d = Driver(
                driver_id=f"EQ.INDEX.{idx}",
                name=idx,
                asset_class=AssetClass.EQUITY,
                shock_mod=shock["mod"],
                shock_sev=shock["sev"],
                is_anchor=(idx == "SPX"),
                sector="Index",
            )
            d.derivation = shock.get("description", "")
            self.drivers[d.driver_id] = d
        
        # Sector ETFs
        for sector, data in EQUITY_SECTOR_SHOCKS.items():
            d = Driver(
                driver_id=f"EQ.SECTOR.{sector}",
                name=sector,
                asset_class=AssetClass.EQUITY,
                shock_mod=data["mod"],
                shock_sev=data["sev"],
                sector=sector,
                beta=data["beta"],
            )
            d.derivation = f"SPX({EQUITY_INDEX_SHOCKS['SPX']['mod']}/{EQUITY_INDEX_SHOCKS['SPX']['sev']}) × β({data['beta']:.2f}) + AI_overlay({data['ai_overlay_mod']}/{data['ai_overlay_sev']})"
            self.drivers[d.driver_id] = d
        
        # Single names from company database
        for ticker, company in COMPANY_DATABASE.items():
            # Check if binding constraint
            if ticker in EQUITY_BINDING_CONSTRAINTS:
                binding = EQUITY_BINDING_CONSTRAINTS[ticker]
                d = Driver(
                    driver_id=f"EQ.US.{ticker}",
                    name=company["name"],
                    asset_class=AssetClass.EQUITY,
                    shock_mod=binding["mod"],
                    shock_sev=binding["sev"],
                    sector=company["sector"],
                    ai_tier=company["ai_tier"],
                    ticker=ticker,
                    beta=company.get("beta", 1.0),
                    is_binding_constraint=True,
                    naic_code=company.get("naic"),
                )
                d.derivation = f"BINDING: {binding.get('rationale', '')}"
            else:
                # Derive from sector
                sector = company["sector"]
                sector_data = EQUITY_SECTOR_SHOCKS.get(sector, {"mod": -22, "sev": -33})
                beta = company.get("beta", 1.0)
                
                # Calculate shock
                base_mod = EQUITY_INDEX_SHOCKS["SPX"]["mod"]
                base_sev = EQUITY_INDEX_SHOCKS["SPX"]["sev"]
                shock_mod = base_mod * beta
                shock_sev = base_sev * beta
                
                d = Driver(
                    driver_id=f"EQ.US.{ticker}",
                    name=company["name"],
                    asset_class=AssetClass.EQUITY,
                    shock_mod=round(shock_mod, 1),
                    shock_sev=round(shock_sev, 1),
                    sector=sector,
                    ai_tier=company["ai_tier"],
                    ticker=ticker,
                    beta=beta,
                    naic_code=company.get("naic"),
                    base_shock_mod=base_mod,
                    base_shock_sev=base_sev,
                    multiplier=beta,
                )
                d.compute_derivation_string()
            
            self.drivers[d.driver_id] = d
    
    def _build_credit_drivers(self):
        """Build all credit drivers."""
        
        # 1. Index
        for idx_name, shock in INDEX_SHOCKS.items():
            bucket = RatingBucket.IG if "IG" in idx_name else RatingBucket.HY
            d = Driver(
                driver_id=f"CR.INDEX.{idx_name}",
                name=idx_name,
                asset_class=AssetClass.CREDIT,
                credit_type=CreditType.INDEX,
                shock_mod=shock["mod"],
                shock_sev=shock["sev"],
                rating=bucket.value,
                is_anchor=True,
            )
            d.derivation = shock.get("derivation", "")
            self.drivers[d.driver_id] = d
        
        # 2. Corporate bond grid: Rating × Sector × Tenor
        key_ratings = ["AAA", "AA", "A", "BBB", "BB", "B", "CCC"]
        key_sectors = list(CREDIT_SECTOR_OVERLAYS.keys())
        key_tenors = ["1Y", "3Y", "5Y", "7Y", "10Y", "30Y"]
        
        for rating in key_ratings:
            for sector in key_sectors:
                for tenor in key_tenors:
                    # Base shock from rating
                    base = CORP_RATING_SHOCKS.get(rating, CORP_RATING_SHOCKS["BBB"])
                    base_mod = base["mod"]
                    base_sev = base["sev"]
                    
                    # Tenor multiplier
                    tenor_data = TENOR_MULTIPLIERS.get(tenor, {"mult": 1.0})
                    tenor_mult = tenor_data["mult"]
                    
                    # Sector overlay
                    sector_overlay = CREDIT_SECTOR_OVERLAYS.get(sector, {"mod": 0, "sev": 0})
                    
                    # AI tier (infer from sector)
                    if sector in ["Semiconductors", "Software"]:
                        ai_tier = AITier.TIER1_DIRECT
                    elif sector == "Technology":
                        ai_tier = AITier.TIER2_INDIRECT
                    elif sector in ["Healthcare", "ConsumerStaples"]:
                        ai_tier = AITier.DEFENSIVE
                    else:
                        ai_tier = AITier.TIER3_MINIMAL
                    
                    ai_overlay = CREDIT_AI_TIER_OVERLAYS.get(ai_tier, {"mod": 0, "sev": 0})
                    
                    # Total shock
                    total_mod = base_mod * tenor_mult + sector_overlay["mod"] + ai_overlay["mod"]
                    total_sev = base_sev * tenor_mult + sector_overlay["sev"] + ai_overlay["sev"]
                    
                    # Ensure monotonicity: |severe| >= |moderate|
                    if abs(total_sev) < abs(total_mod):
                        total_sev = total_mod * (base_sev / base_mod) if base_mod != 0 else total_mod * 1.5
                    
                    bucket = "IG" if rating in ["AAA", "AA", "A", "BBB"] else "HY"
                    
                    d = Driver(
                        driver_id=f"CR.CORP.{bucket}.{rating}.{sector}.{tenor}",
                        name=f"{rating} {sector} {tenor}",
                        asset_class=AssetClass.CREDIT,
                        credit_type=CreditType.CORPORATE,
                        shock_mod=round(total_mod, 1),
                        shock_sev=round(total_sev, 1),
                        rating=rating,
                        sector=sector,
                        tenor=tenor,
                        ai_tier=ai_tier,
                        base_shock_mod=base_mod,
                        base_shock_sev=base_sev,
                        multiplier=tenor_mult,
                        overlay_mod=sector_overlay["mod"] + ai_overlay["mod"],
                        overlay_sev=sector_overlay["sev"] + ai_overlay["sev"],
                    )
                    d.compute_derivation_string()
                    self.drivers[d.driver_id] = d
        
        # 3. CDS for all companies
        for ticker, company in COMPANY_DATABASE.items():
            rating = company.get("rating", "BBB")
            cds_base = CDS_RATING_SHOCKS.get(rating, CDS_RATING_SHOCKS["BBB"])
            ai_overlay = CREDIT_AI_TIER_OVERLAYS.get(company["ai_tier"], {"mod": 0, "sev": 0})
            sector_overlay = CREDIT_SECTOR_OVERLAYS.get(company["sector"], {"mod": 0, "sev": 0})
            
            total_mod = cds_base["mod"] + sector_overlay["mod"] * 0.5 + ai_overlay["mod"] * 0.5  # CDS less sensitive
            total_sev = cds_base["sev"] + sector_overlay["sev"] * 0.5 + ai_overlay["sev"] * 0.5
            
            d = Driver(
                driver_id=f"CDS.{ticker}.5Y",
                name=f"{company['name']} CDS 5Y",
                asset_class=AssetClass.CREDIT,
                credit_type=CreditType.CDS,
                shock_mod=round(total_mod, 1),
                shock_sev=round(total_sev, 1),
                rating=rating,
                sector=company["sector"],
                ai_tier=company["ai_tier"],
                ticker=ticker,
                tenor="5Y",
                naic_code=company.get("naic"),
                base_shock_mod=cds_base["mod"],
                base_shock_sev=cds_base["sev"],
            )
            d.compute_derivation_string()
            self.drivers[d.driver_id] = d
    
    def _build_commodity_drivers(self):
        """Build commodity drivers."""
        for commodity, data in COMMODITY_SHOCKS.items():
            d = Driver(
                driver_id=f"CO.{data['type'].upper()}.{commodity}",
                name=commodity,
                asset_class=AssetClass.COMMODITY,
                shock_mod=data["mod"],
                shock_sev=data["sev"],
                commodity_type=data["type"],
                is_anchor=(commodity in ["WTI", "XAU"]),
            )
            d.derivation = data.get("derivation", "")
            self.drivers[d.driver_id] = d
    
    def _build_rates_drivers(self):
        """Build rates drivers."""
        for rate, data in RATES_SHOCKS.items():
            d = Driver(
                driver_id=f"RT.{rate}",
                name=rate,
                asset_class=AssetClass.RATES,
                shock_mod=data["mod"],
                shock_sev=data["sev"],
                is_anchor=(rate == "UST.10Y"),
            )
            d.derivation = data.get("derivation", "")
            self.drivers[d.driver_id] = d
    
    def _build_fx_drivers(self):
        """Build FX drivers."""
        for pair, data in FX_SHOCKS.items():
            d = Driver(
                driver_id=f"FX.{pair}",
                name=pair,
                asset_class=AssetClass.FX,
                shock_mod=data["mod"],
                shock_sev=data["sev"],
                is_anchor=(pair == "DXY"),
            )
            d.derivation = data.get("derivation", "")
            self.drivers[d.driver_id] = d
    
    def _build_vol_drivers(self):
        """Build volatility drivers."""
        for vol_id, data in VOL_SHOCKS.items():
            shock_mod = data.get("mod_chg", data.get("mod", 0))
            shock_sev = data.get("sev_chg", data.get("sev", 0))
            
            d = Driver(
                driver_id=f"VOL.{vol_id}",
                name=vol_id,
                asset_class=AssetClass.VOLATILITY,
                shock_mod=shock_mod,
                shock_sev=shock_sev,
                is_anchor=(vol_id == "VIX"),
            )
            if "mod_level" in data:
                d.stressed_level_mod = data["mod_level"]
                d.stressed_level_sev = data["sev_level"]
            d.derivation = data.get("derivation", "")
            self.drivers[d.driver_id] = d
    
    # -------------------------------------------------------------------------
    # VALIDATION
    # -------------------------------------------------------------------------
    
    def _run_validation(self):
        """Run all validation checks."""
        self._validate_monotonicity()
        self._validate_cross_asset_consistency()
        self._validate_historical_bounds()
        self._validate_sign_consistency()
        self._validate_ai_tier_ordering()
    
    def _validate_monotonicity(self):
        """Check that |severe| >= |moderate| for all drivers."""
        for d in self.drivers.values():
            if abs(d.shock_sev) < abs(d.shock_mod) - 0.01:  # tolerance for rounding
                self.validation_results.append(ValidationResult(
                    check_name="monotonicity",
                    driver_id=d.driver_id,
                    severity=ValidationSeverity.ERROR,
                    message=f"Severe shock ({d.shock_sev}) < Moderate shock ({d.shock_mod})",
                    expected=f"|{d.shock_mod}| <= |severe|",
                    actual=f"|{d.shock_sev}|",
                ))
    
    def _validate_cross_asset_consistency(self):
        """Check cross-asset relationships."""
        # VIX should rise when SPX falls
        spx = self.drivers.get("EQ.INDEX.SPX")
        vix = self.drivers.get("VOL.VIX")
        if spx and vix:
            if spx.shock_mod < 0 and vix.shock_mod <= 0:
                self.validation_results.append(ValidationResult(
                    check_name="cross_asset",
                    driver_id="VIX vs SPX",
                    severity=ValidationSeverity.ERROR,
                    message="VIX should rise when SPX falls",
                    expected="VIX_shock > 0",
                    actual=f"VIX_shock = {vix.shock_mod}",
                ))
        
        # Credit should widen when equity falls
        ig_idx = self.drivers.get("CR.INDEX.CDX.NA.IG")
        if spx and ig_idx:
            if spx.shock_mod < 0 and ig_idx.shock_mod <= 0:
                self.validation_results.append(ValidationResult(
                    check_name="cross_asset",
                    driver_id="Credit vs Equity",
                    severity=ValidationSeverity.ERROR,
                    message="Credit spreads should widen when equity falls",
                    expected="IG_OAS_shock > 0",
                    actual=f"IG_OAS_shock = {ig_idx.shock_mod}",
                ))
        
        # Gold should rise in risk-off
        gold = self.drivers.get("CO.PRECIOUS.XAU")
        if spx and gold:
            if spx.shock_sev < -20 and gold.shock_sev <= 0:
                self.validation_results.append(ValidationResult(
                    check_name="cross_asset",
                    driver_id="Gold vs SPX",
                    severity=ValidationSeverity.WARNING,
                    message="Gold typically rises as safe haven in severe equity selloff",
                    expected="XAU_shock > 0",
                    actual=f"XAU_shock = {gold.shock_sev}",
                ))
        
        # USD should strengthen in risk-off
        dxy = self.drivers.get("FX.DXY")
        if spx and dxy:
            if spx.shock_sev < -20 and dxy.shock_sev <= 0:
                self.validation_results.append(ValidationResult(
                    check_name="cross_asset",
                    driver_id="DXY vs SPX",
                    severity=ValidationSeverity.WARNING,
                    message="USD typically strengthens in risk-off",
                    expected="DXY_shock > 0",
                    actual=f"DXY_shock = {dxy.shock_sev}",
                ))
    
    def _validate_historical_bounds(self):
        """Check shocks against historical extremes."""
        historical_bounds = {
            "EQ.INDEX.SPX": {"min_mod": -30, "max_mod": -10, "min_sev": -50, "max_sev": -20,
                            "reference": "COVID=-34%, GFC=-57%, 2022=-25%"},
            "VOL.VIX": {"min_mod": 25, "max_mod": 50, "min_sev": 40, "max_sev": 90,
                       "reference": "COVID=82, GFC=80, 2022=36"},
            "CR.INDEX.CDX.NA.IG": {"min_mod": 10, "max_mod": 30, "min_sev": 80, "max_sev": 250,
                                   "reference": "COVID=150bp, GFC=300bp, 2022=80bp"},
            "CR.INDEX.CDX.NA.HY": {"min_mod": 30, "max_mod": 80, "min_sev": 200, "max_sev": 700,
                                   "reference": "COVID=600bp, GFC=1600bp, 2022=280bp"},
        }
        
        for driver_id, bounds in historical_bounds.items():
            d = self.drivers.get(driver_id)
            if d:
                # Check moderate
                if not (bounds["min_mod"] <= abs(d.shock_mod) <= bounds["max_mod"]):
                    self.validation_results.append(ValidationResult(
                        check_name="historical_bounds",
                        driver_id=driver_id,
                        severity=ValidationSeverity.WARNING,
                        message=f"Moderate shock outside typical range",
                        expected=f"{bounds['min_mod']} to {bounds['max_mod']} ({bounds['reference']})",
                        actual=str(d.shock_mod),
                    ))
                
                # Check severe
                if not (bounds["min_sev"] <= abs(d.shock_sev) <= bounds["max_sev"]):
                    self.validation_results.append(ValidationResult(
                        check_name="historical_bounds",
                        driver_id=driver_id,
                        severity=ValidationSeverity.WARNING,
                        message=f"Severe shock outside typical range",
                        expected=f"{bounds['min_sev']} to {bounds['max_sev']} ({bounds['reference']})",
                        actual=str(d.shock_sev),
                    ))
    
    def _validate_sign_consistency(self):
        """Check that shock signs are consistent with risk-off scenario."""
        expected_signs = {
            # Equity should fall (negative)
            "EQ.INDEX.SPX": "negative",
            "EQ.INDEX.NDX": "negative",
            # Credit should widen (positive)
            "CR.INDEX.CDX.NA.IG": "positive",
            "CR.INDEX.CDX.NA.HY": "positive",
            # Vol should rise (positive)
            "VOL.VIX": "positive",
            # Rates should fall (negative = rally)
            "RT.UST.10Y": "negative",
            # USD should strengthen (positive)
            "FX.DXY": "positive",
            # Gold should rise (positive)
            "CO.PRECIOUS.XAU": "positive",
            # Oil should fall (negative)
            "CO.ENERGY.WTI": "negative",
        }
        
        for driver_id, expected in expected_signs.items():
            d = self.drivers.get(driver_id)
            if d:
                actual_sign = "positive" if d.shock_sev > 0 else "negative" if d.shock_sev < 0 else "zero"
                if actual_sign != expected:
                    self.validation_results.append(ValidationResult(
                        check_name="sign_consistency",
                        driver_id=driver_id,
                        severity=ValidationSeverity.ERROR,
                        message=f"Shock sign inconsistent with risk-off scenario",
                        expected=expected,
                        actual=f"{actual_sign} ({d.shock_sev})",
                    ))
    
    def _validate_ai_tier_ordering(self):
        """Check that AI Tier 1 has larger shocks than Tier 3 for same sector."""
        tier1_drivers = [d for d in self.drivers.values() 
                        if d.ai_tier == AITier.TIER1_DIRECT and d.asset_class == AssetClass.EQUITY]
        tier3_drivers = [d for d in self.drivers.values() 
                        if d.ai_tier == AITier.TIER3_MINIMAL and d.asset_class == AssetClass.EQUITY]
        
        if tier1_drivers and tier3_drivers:
            avg_tier1 = sum(abs(d.shock_sev) for d in tier1_drivers) / len(tier1_drivers)
            avg_tier3 = sum(abs(d.shock_sev) for d in tier3_drivers) / len(tier3_drivers)
            
            if avg_tier1 <= avg_tier3:
                self.validation_results.append(ValidationResult(
                    check_name="ai_tier_ordering",
                    driver_id="AI_TIER_COMPARISON",
                    severity=ValidationSeverity.ERROR,
                    message="Tier 1 (direct AI) should have larger shocks than Tier 3 (minimal)",
                    expected=f"Avg Tier 1 > Avg Tier 3",
                    actual=f"Tier1={avg_tier1:.1f}%, Tier3={avg_tier3:.1f}%",
                ))
    
    # -------------------------------------------------------------------------
    # EXPORT METHODS
    # -------------------------------------------------------------------------
    
    def summary_stats(self) -> Dict:
        """Generate summary statistics."""
        stats = {
            "total_drivers": len(self.drivers),
            "by_asset_class": {},
            "by_ai_tier": {},
            "anchors": [],
            "binding_constraints": [],
            "validation_errors": sum(1 for v in self.validation_results if v.severity == ValidationSeverity.ERROR),
            "validation_warnings": sum(1 for v in self.validation_results if v.severity == ValidationSeverity.WARNING),
        }
        
        for d in self.drivers.values():
            # By asset class
            ac = d.asset_class.value
            stats["by_asset_class"][ac] = stats["by_asset_class"].get(ac, 0) + 1
            
            # By AI tier
            tier = d.ai_tier.value
            stats["by_ai_tier"][tier] = stats["by_ai_tier"].get(tier, 0) + 1
            
            # Anchors
            if d.is_anchor:
                stats["anchors"].append(d.driver_id)
            
            # Binding constraints
            if d.is_binding_constraint:
                stats["binding_constraints"].append(d.driver_id)
        
        return stats
    
    def to_dict_list(self) -> List[Dict]:
        """Export all drivers to list of dicts."""
        output = []
        for d in sorted(self.drivers.values(), key=lambda x: (x.asset_class.value, x.driver_id)):
            output.append({
                "Driver_ID": d.driver_id,
                "Name": d.name,
                "Asset_Class": d.asset_class.value,
                "Sector": d.sector,
                "AI_Tier": d.ai_tier.value,
                "NAIC_Code": d.naic_code,
                "Rating": d.rating,
                "Tenor": d.tenor,
                "Credit_Type": d.credit_type.value if d.credit_type else None,
                "Beta": d.beta if d.beta != 1.0 else None,
                "Ticker": d.ticker,
                "Is_Anchor": d.is_anchor,
                "Is_Binding": d.is_binding_constraint,
                "Shock_Moderate": round(d.shock_mod, 2),
                "Shock_Severe": round(d.shock_sev, 2),
                "Derivation": d.derivation,
            })
        return output


# =============================================================================
# SECTION 8: EXCEL EXPORT
# =============================================================================

def export_to_excel(engine: UnifiedStressEngine, filepath: str):
    """Export complete scenario to Excel with full documentation."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    anchor_fill = PatternFill('solid', fgColor='90EE90')
    binding_fill = PatternFill('solid', fgColor='FFD700')
    tier1_fill = PatternFill('solid', fgColor='FF6B6B')
    tier2_fill = PatternFill('solid', fgColor='FFB347')
    tier3_fill = PatternFill('solid', fgColor='FFFFFF')
    defensive_fill = PatternFill('solid', fgColor='90EE90')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
    
    # =========================================================================
    # SHEET 1: Executive Summary
    # =========================================================================
    ws = wb.active
    ws.title = "Executive Summary"
    
    stats = engine.summary_stats()
    
    summary_rows = [
        ["AI BUBBLE CORRECTION STRESS SCENARIO", ""],
        [f"Generated: {engine.generated_at.strftime('%Y-%m-%d %H:%M')}", ""],
        ["", ""],
        ["SCENARIO OVERVIEW", ""],
        ["Moderate Scenario:", "10-day horizon, orderly correction"],
        ["Severe Scenario:", "60-day horizon, disorderly unwind with contagion"],
        ["", ""],
        ["ANCHOR SHOCKS", ""],
        ["SPX:", f"{EQUITY_INDEX_SHOCKS['SPX']['mod']}% / {EQUITY_INDEX_SHOCKS['SPX']['sev']}%"],
        ["VIX:", f"{VOL_SHOCKS['VIX']['mod_level']} / {VOL_SHOCKS['VIX']['sev_level']} (level)"],
        ["IG OAS:", f"+{INDEX_SHOCKS['CDX.NA.IG']['mod']}bp / +{INDEX_SHOCKS['CDX.NA.IG']['sev']}bp"],
        ["HY OAS:", f"+{INDEX_SHOCKS['CDX.NA.HY']['mod']}bp / +{INDEX_SHOCKS['CDX.NA.HY']['sev']}bp"],
        ["UST 10Y:", f"{RATES_SHOCKS['UST.10Y']['mod']}bp / {RATES_SHOCKS['UST.10Y']['sev']}bp"],
        ["DXY:", f"+{FX_SHOCKS['DXY']['mod']}% / +{FX_SHOCKS['DXY']['sev']}%"],
        ["WTI:", f"{COMMODITY_SHOCKS['WTI']['mod']}% / {COMMODITY_SHOCKS['WTI']['sev']}%"],
        ["Gold:", f"+{COMMODITY_SHOCKS['XAU']['mod']}% / +{COMMODITY_SHOCKS['XAU']['sev']}%"],
        ["", ""],
        ["DRIVER COUNTS", ""],
        ["Total Drivers:", stats["total_drivers"]],
    ]
    
    for ac, count in stats["by_asset_class"].items():
        summary_rows.append([f"  {ac}:", count])
    
    summary_rows.extend([
        ["", ""],
        ["BY AI TIER", ""],
    ])
    for tier, count in stats["by_ai_tier"].items():
        summary_rows.append([f"  {tier}:", count])
    
    summary_rows.extend([
        ["", ""],
        ["VALIDATION", ""],
        ["Errors:", stats["validation_errors"]],
        ["Warnings:", stats["validation_warnings"]],
    ])
    
    for row_idx, (label, value) in enumerate(summary_rows, 1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label and not label.startswith(" ") and ":" not in label:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    
    # =========================================================================
    # SHEET 2: All Drivers
    # =========================================================================
    ws = wb.create_sheet("All Drivers")
    
    headers = [
        "Driver_ID", "Name", "Asset_Class", "Sector", "AI_Tier", "Rating", "Tenor",
        "Shock_Mod", "Shock_Sev", "Is_Anchor", "Is_Binding", "Derivation"
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    
    tier_fills = {
        AITier.TIER1_DIRECT: tier1_fill,
        AITier.TIER2_INDIRECT: tier2_fill,
        AITier.TIER3_MINIMAL: tier3_fill,
        AITier.DEFENSIVE: defensive_fill,
    }
    
    for d in sorted(engine.drivers.values(), key=lambda x: (x.asset_class.value, x.driver_id)):
        row = [
            d.driver_id, d.name, d.asset_class.value, d.sector, d.ai_tier.value,
            d.rating, d.tenor, d.shock_mod, d.shock_sev,
            "Yes" if d.is_anchor else "", "Yes" if d.is_binding_constraint else "",
            d.derivation
        ]
        ws.append(row)
        
        row_num = ws.max_row
        fill = tier_fills.get(d.ai_tier, tier3_fill)
        if d.is_anchor:
            fill = anchor_fill
        elif d.is_binding_constraint:
            fill = binding_fill
        
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
            ws.cell(row=row_num, column=col).border = thin_border
    
    widths = [30, 30, 12, 18, 15, 8, 6, 10, 10, 10, 10, 60]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # =========================================================================
    # SHEET 3: Equity Drivers
    # =========================================================================
    ws = wb.create_sheet("Equity")
    
    eq_headers = ["Driver_ID", "Name", "Sector", "AI_Tier", "Beta", "Shock_Mod_%", "Shock_Sev_%", "Derivation"]
    ws.append(eq_headers)
    style_header(ws, len(eq_headers))
    
    for d in sorted(engine.drivers.values(), key=lambda x: x.driver_id):
        if d.asset_class != AssetClass.EQUITY:
            continue
        ws.append([
            d.driver_id, d.name, d.sector, d.ai_tier.value, d.beta,
            d.shock_mod, d.shock_sev, d.derivation
        ])
        
        row_num = ws.max_row
        fill = tier_fills.get(d.ai_tier, tier3_fill)
        if d.is_anchor:
            fill = anchor_fill
        elif d.is_binding_constraint:
            fill = binding_fill
        for col in range(1, len(eq_headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
    
    for col_idx, width in enumerate([28, 28, 18, 15, 6, 12, 12, 50], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # =========================================================================
    # SHEET 4: Credit Drivers
    # =========================================================================
    ws = wb.create_sheet("Credit")
    
    cr_headers = ["Driver_ID", "Name", "Type", "Rating", "Sector", "Tenor", "AI_Tier",
                  "Base_Mod", "Base_Sev", "Mult", "Overlay", "Total_Mod", "Total_Sev"]
    ws.append(cr_headers)
    style_header(ws, len(cr_headers))
    
    for d in sorted(engine.drivers.values(), key=lambda x: (x.credit_type.value if x.credit_type else "", x.rating or "", x.driver_id)):
        if d.asset_class != AssetClass.CREDIT:
            continue
        ws.append([
            d.driver_id, d.name, d.credit_type.value if d.credit_type else "",
            d.rating, d.sector, d.tenor, d.ai_tier.value,
            d.base_shock_mod, d.base_shock_sev, d.multiplier,
            f"{d.overlay_mod:+.0f}/{d.overlay_sev:+.0f}" if d.overlay_mod or d.overlay_sev else "",
            d.shock_mod, d.shock_sev
        ])
        
        row_num = ws.max_row
        fill = tier_fills.get(d.ai_tier, tier3_fill)
        if d.is_anchor:
            fill = anchor_fill
        for col in range(1, len(cr_headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
    
    # =========================================================================
    # SHEET 5: Shock Grids (Reference)
    # =========================================================================
    ws = wb.create_sheet("Shock Grids")
    
    row = 1
    ws.cell(row=row, column=1, value="CORPORATE BOND RATING GRID").font = Font(bold=True, size=14)
    row += 1
    ws.append(["Rating", "Mod_bp", "Sev_bp", "Derivation"])
    style_header(ws, 4)
    for rating, data in CORP_RATING_SHOCKS.items():
        ws.append([rating, data["mod"], data["sev"], data["derivation"]])
    
    row = ws.max_row + 3
    ws.cell(row=row, column=1, value="CDS RATING GRID").font = Font(bold=True, size=14)
    row += 1
    for col_idx, header in enumerate(["Rating", "Mod_bp", "Sev_bp", "Derivation"], 1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = header_font
        ws.cell(row=row, column=col_idx).fill = header_fill
    row += 1
    for rating, data in CDS_RATING_SHOCKS.items():
        ws.cell(row=row, column=1, value=rating)
        ws.cell(row=row, column=2, value=data["mod"])
        ws.cell(row=row, column=3, value=data["sev"])
        ws.cell(row=row, column=4, value=data["derivation"])
        row += 1
    
    row += 2
    ws.cell(row=row, column=1, value="TENOR MULTIPLIERS").font = Font(bold=True, size=14)
    row += 1
    for col_idx, header in enumerate(["Tenor", "Multiplier", "Derivation"], 1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = header_font
        ws.cell(row=row, column=col_idx).fill = header_fill
    row += 1
    for tenor, data in TENOR_MULTIPLIERS.items():
        ws.cell(row=row, column=1, value=tenor)
        ws.cell(row=row, column=2, value=data["mult"])
        ws.cell(row=row, column=3, value=data["derivation"])
        row += 1
    
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = [10, 10, 10, 50][col-1]
    
    # =========================================================================
    # SHEET 6: AI Tier Classification
    # =========================================================================
    ws = wb.create_sheet("AI Tier Classification")
    
    tier_headers = ["AI_Tier", "EQ_Impact", "CR_Overlay_Mod", "CR_Overlay_Sev", "Description", "Example_Companies"]
    ws.append(tier_headers)
    style_header(ws, len(tier_headers))
    
    tier_examples = {
        AITier.TIER1_DIRECT: "NVDA, AMD, AVGO, PLTR, SNOW, TSM, ASML",
        AITier.TIER2_INDIRECT: "MSFT, GOOGL, AMZN, META, VST, CEG, EQIX, TSLA",
        AITier.TIER3_MINIMAL: "AAPL, JPM, XOM, BA, CAT",
        AITier.DEFENSIVE: "JNJ, PG, KO, WMT, PFE",
    }
    
    for tier, overlay in CREDIT_AI_TIER_OVERLAYS.items():
        ws.append([
            tier.value,
            "Epicenter" if tier == AITier.TIER1_DIRECT else "Elevated" if tier == AITier.TIER2_INDIRECT else "Market" if tier == AITier.TIER3_MINIMAL else "Outperform",
            overlay["mod"],
            overlay["sev"],
            overlay["derivation"],
            tier_examples.get(tier, ""),
        ])
        row_num = ws.max_row
        fill = tier_fills.get(tier, tier3_fill)
        for col in range(1, len(tier_headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
    
    # =========================================================================
    # SHEET 7: Validation Results
    # =========================================================================
    ws = wb.create_sheet("Validation")
    
    val_headers = ["Check", "Driver_ID", "Severity", "Message", "Expected", "Actual"]
    ws.append(val_headers)
    style_header(ws, len(val_headers))
    
    error_fill = PatternFill('solid', fgColor='FF6B6B')
    warning_fill = PatternFill('solid', fgColor='FFD700')
    
    for v in engine.validation_results:
        ws.append([v.check_name, v.driver_id, v.severity.value, v.message, str(v.expected), str(v.actual)])
        row_num = ws.max_row
        fill = error_fill if v.severity == ValidationSeverity.ERROR else warning_fill
        for col in range(1, len(val_headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
    
    if not engine.validation_results:
        ws.append(["", "", "", "All validation checks passed!", "", ""])
        ws.cell(row=2, column=4).fill = anchor_fill
    
    # =========================================================================
    # SHEET 8: Company Database
    # =========================================================================
    ws = wb.create_sheet("Company Database")
    
    co_headers = ["Ticker", "Name", "Sector", "Rating", "AI_Tier", "Beta", "AI_Revenue_%", "NAIC", "Description"]
    ws.append(co_headers)
    style_header(ws, len(co_headers))
    
    for ticker, company in sorted(COMPANY_DATABASE.items(), key=lambda x: (x[1]["ai_tier"].value, x[0])):
        ws.append([
            ticker, company["name"], company["sector"], company["rating"],
            company["ai_tier"].value, company.get("beta", 1.0),
            company.get("ai_revenue_pct", 0), company.get("naic", ""),
            company.get("description", "")
        ])
        row_num = ws.max_row
        fill = tier_fills.get(company["ai_tier"], tier3_fill)
        for col in range(1, len(co_headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
    
    # =========================================================================
    # SHEET 9: Historical Calibration
    # =========================================================================
    ws = wb.create_sheet("Historical Calibration")
    
    hist_headers = ["Episode", "SPX_Drawdown", "NDX_Drawdown", "VIX_Peak", "IG_Widening", "HY_Widening", "Duration"]
    ws.append(hist_headers)
    style_header(ws, len(hist_headers))
    
    for episode, data in HISTORICAL_EPISODES.items():
        ws.append([
            data["description"],
            f"{data['spx_drawdown']}%",
            f"{data['ndx_drawdown']}%",
            data["vix_peak"],
            f"+{data['ig_spread_widening']}bp",
            f"+{data['hy_spread_widening']}bp",
            f"{data.get('duration_months', data.get('duration_days', 'N/A'))} {'months' if 'duration_months' in data else 'days'}",
        ])
    
    row = ws.max_row + 3
    ws.cell(row=row, column=1, value="OUR SCENARIO CALIBRATION").font = Font(bold=True, size=14)
    row += 1
    for scenario, data in SCENARIO_CALIBRATION.items():
        row += 1
        ws.cell(row=row, column=1, value=scenario.upper()).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value=f"  Horizon: {data['horizon_days']} days")
        row += 1
        ws.cell(row=row, column=1, value=f"  Historical Analog: {data['historical_analog']}")
        row += 1
        ws.cell(row=row, column=1, value=f"  SPX Target: {data['spx_target']}%")
        row += 1
        ws.cell(row=row, column=1, value=f"  VIX Target: {data['vix_target']}")
    
    # Save
    wb.save(filepath)
    print(f"Exported to: {filepath}")


# =============================================================================
# SECTION 9: MAIN EXECUTION
# =============================================================================

def main():
    """Build and export complete stress scenario."""
    
    # Create engine
    engine = UnifiedStressEngine()
    
    # Build full scenario
    engine.build_full_scenario()
    
    # Export to Excel
    export_to_excel(engine, "/home/claude/ai_bubble_stress_complete.xlsx")
    
    # Export to JSON
    with open("/home/claude/ai_bubble_stress_complete.json", 'w') as f:
        json.dump({
            "scenario_name": engine.scenario_name,
            "generated_at": engine.generated_at.isoformat(),
            "summary": engine.summary_stats(),
            "drivers": engine.to_dict_list(),
            "validation": [
                {
                    "check": v.check_name,
                    "driver_id": v.driver_id,
                    "severity": v.severity.value,
                    "message": v.message,
                }
                for v in engine.validation_results
            ],
        }, f, indent=2)
    print("Exported JSON: /home/claude/ai_bubble_stress_complete.json")
    
    # Print summary
    stats = engine.summary_stats()
    print(f"\n{'=' * 70}")
    print("SUMMARY")
    print(f"{'=' * 70}")
    print(f"Total drivers: {stats['total_drivers']}")
    print(f"By asset class: {stats['by_asset_class']}")
    print(f"Anchors: {len(stats['anchors'])}")
    print(f"Binding constraints: {len(stats['binding_constraints'])}")
    print(f"Validation errors: {stats['validation_errors']}")
    print(f"Validation warnings: {stats['validation_warnings']}")
    
    return engine


if __name__ == "__main__":
    engine = main()
