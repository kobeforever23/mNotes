#!/usr/bin/env python3
"""
================================================================================
EQUITY STRESS ANALYZER - AI BUBBLE SCENARIO VALIDATION
================================================================================
Single-file tool with GUI for PyCharm

Requirements: pip install yfinance openpyxl pandas numpy

Just run this file in PyCharm and the GUI will open.
================================================================================
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import os
import warnings
warnings.filterwarnings('ignore')

# Check for optional packages
try:
    import yfinance as yf
    YF_AVAILABLE = True
except ImportError:
    YF_AVAILABLE = False
    print("Note: yfinance not installed. Install with: pip install yfinance")

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    TK_AVAILABLE = True
except ImportError:
    TK_AVAILABLE = False
    print("Note: tkinter not available. Using command-line mode.")

# =============================================================================
# CONFIGURATION - YOUR SCENARIO PARAMETERS
# =============================================================================

SCENARIO_CONFIG = {
    "severe": {
        "horizon_days": 60,
        "single_name_shock": (-0.50, -0.40),  # -50% to -40%
        "utilities_shock": (-0.30, -0.15),     # -30% to -15%
    },
    "moderate": {
        "horizon_days": 10,
        "single_name_shock": (-0.25, -0.15),  # -25% to -15%
        "utilities_shock": (-0.15, -0.08),     # -15% to -8%
    }
}

HISTORICAL_PERIODS = {
    "2000_dotcom": {
        "name": "2000 Dot-Com Crash",
        "start": "2000-03-01",
        "end": "2002-10-31",
    },
    "2018_tech": {
        "name": "2018 Q4 Tech Correction",
        "start": "2018-09-01",
        "end": "2018-12-31",
    },
    "2022_duration": {
        "name": "2022 Duration Derating",
        "start": "2022-01-01",
        "end": "2022-12-31",
    },
    "2025_tariffs": {
        "name": "2025 Tariff Volatility",
        "start": "2025-01-20",
        "end": "2025-04-30",
    }
}

DEFAULT_TICKERS = {
    "ai_tech": ["NVDA", "AMD", "MSFT", "GOOGL", "META", "AMZN", "AAPL", "AVGO", "CRM", "NOW"],
    "utilities": ["NEE", "DUK", "SO", "D", "AEP", "XEL"],
    "benchmarks": ["SPY", "QQQ"]
}


# =============================================================================
# CORE ANALYSIS ENGINE
# =============================================================================

class StressAnalyzer:
    """Core analysis engine for beta and stress testing"""
    
    def __init__(self, prices: pd.DataFrame):
        self.prices = prices
        self.returns = np.log(prices / prices.shift(1)).dropna()
    
    def calculate_beta(self, ticker: str, benchmark: str = "SPY") -> float:
        """Calculate beta of ticker vs benchmark"""
        if ticker not in self.returns.columns or benchmark not in self.returns.columns:
            return np.nan
        
        cov = np.cov(self.returns[ticker].dropna(), self.returns[benchmark].dropna())[0, 1]
        var = np.var(self.returns[benchmark].dropna())
        return cov / var if var != 0 else np.nan
    
    def calculate_all_betas(self, tickers: List[str]) -> pd.DataFrame:
        """Calculate betas for all tickers"""
        results = []
        for ticker in tickers:
            if ticker not in self.prices.columns:
                continue
            
            beta_spy = self.calculate_beta(ticker, "SPY") if "SPY" in self.prices.columns else np.nan
            beta_qqq = self.calculate_beta(ticker, "QQQ") if "QQQ" in self.prices.columns else np.nan
            
            ret = self.returns[ticker]
            results.append({
                "Ticker": ticker,
                "Beta_SPY": round(beta_spy, 3) if not np.isnan(beta_spy) else np.nan,
                "Beta_QQQ": round(beta_qqq, 3) if not np.isnan(beta_qqq) else np.nan,
                "Ann_Vol": round(ret.std() * np.sqrt(252), 3),
                "Ann_Return": round(ret.mean() * 252, 3)
            })
        
        return pd.DataFrame(results)
    
    def analyze_period(self, ticker: str, start: str, end: str) -> Dict:
        """Analyze a specific period for a ticker"""
        if ticker not in self.prices.columns:
            return {}
        
        try:
            price = self.prices[ticker][start:end].dropna()
            if len(price) < 10:
                return {}
            
            # Max drawdown
            running_max = price.expanding().max()
            drawdown = (price - running_max) / running_max
            max_dd = drawdown.min()
            
            # Worst N-day returns
            worst_10d = price.pct_change(10).min() if len(price) > 10 else np.nan
            worst_60d = price.pct_change(60).min() if len(price) > 60 else np.nan
            
            # Total period return
            total_return = (price.iloc[-1] / price.iloc[0]) - 1
            
            return {
                "max_drawdown": max_dd,
                "worst_10d": worst_10d,
                "worst_60d": worst_60d,
                "total_return": total_return
            }
        except:
            return {}
    
    def run_full_analysis(self, ai_tickers: List[str], util_tickers: List[str]) -> Dict:
        """Run complete analysis and return all results"""
        
        all_tickers = [t for t in ai_tickers + util_tickers if t in self.prices.columns]
        
        # Beta analysis
        betas = self.calculate_all_betas(all_tickers)
        
        # Historical period analysis
        period_results = []
        for period_key, period_config in HISTORICAL_PERIODS.items():
            for ticker in all_tickers:
                result = self.analyze_period(ticker, period_config["start"], period_config["end"])
                if result:
                    period_results.append({
                        "Period": period_config["name"],
                        "Ticker": ticker,
                        "Type": "AI/Tech" if ticker in ai_tickers else "Utility",
                        "Max_Drawdown": round(result["max_drawdown"] * 100, 1),
                        "Worst_10d": round(result["worst_10d"] * 100, 1) if not np.isnan(result["worst_10d"]) else np.nan,
                        "Worst_60d": round(result["worst_60d"] * 100, 1) if not np.isnan(result["worst_60d"]) else np.nan,
                        "Total_Return": round(result["total_return"] * 100, 1)
                    })
        
        period_df = pd.DataFrame(period_results)
        
        # Summary by period and type
        summary_results = []
        for period_key, period_config in HISTORICAL_PERIODS.items():
            period_data = period_df[period_df["Period"] == period_config["name"]]
            
            ai_data = period_data[period_data["Type"] == "AI/Tech"]
            util_data = period_data[period_data["Type"] == "Utility"]
            
            summary_results.append({
                "Period": period_config["name"],
                "AI_Avg_MaxDD": round(ai_data["Max_Drawdown"].mean(), 1) if len(ai_data) > 0 else np.nan,
                "AI_Worst_MaxDD": round(ai_data["Max_Drawdown"].min(), 1) if len(ai_data) > 0 else np.nan,
                "AI_Worst_60d": round(ai_data["Worst_60d"].min(), 1) if len(ai_data) > 0 else np.nan,
                "AI_Worst_10d": round(ai_data["Worst_10d"].min(), 1) if len(ai_data) > 0 else np.nan,
                "Util_Avg_MaxDD": round(util_data["Max_Drawdown"].mean(), 1) if len(util_data) > 0 else np.nan,
                "Util_Worst_MaxDD": round(util_data["Max_Drawdown"].min(), 1) if len(util_data) > 0 else np.nan,
                "Util_Worst_60d": round(util_data["Worst_60d"].min(), 1) if len(util_data) > 0 else np.nan,
                "Util_Worst_10d": round(util_data["Worst_10d"].min(), 1) if len(util_data) > 0 else np.nan,
            })
        
        summary_df = pd.DataFrame(summary_results)
        
        # Shock validation
        ai_period_data = period_df[period_df["Type"] == "AI/Tech"]
        util_period_data = period_df[period_df["Type"] == "Utility"]
        
        ai_worst_60d = ai_period_data["Worst_60d"].min() / 100 if len(ai_period_data) > 0 else np.nan
        ai_worst_10d = ai_period_data["Worst_10d"].min() / 100 if len(ai_period_data) > 0 else np.nan
        util_worst_60d = util_period_data["Worst_60d"].min() / 100 if len(util_period_data) > 0 else np.nan
        util_worst_10d = util_period_data["Worst_10d"].min() / 100 if len(util_period_data) > 0 else np.nan
        
        validation = [
            {
                "Scenario": "Severe 60-day",
                "Asset_Class": "AI/Tech",
                "Proposed_Shock": f"{SCENARIO_CONFIG['severe']['single_name_shock'][0]*100:.0f}% to {SCENARIO_CONFIG['severe']['single_name_shock'][1]*100:.0f}%",
                "Historical_Worst": f"{ai_worst_60d*100:.1f}%" if not np.isnan(ai_worst_60d) else "N/A",
                "Status": "✓ SUPPORTED" if ai_worst_60d <= SCENARIO_CONFIG['severe']['single_name_shock'][0] else "⚠ AGGRESSIVE"
            },
            {
                "Scenario": "Severe 60-day",
                "Asset_Class": "Utilities",
                "Proposed_Shock": f"{SCENARIO_CONFIG['severe']['utilities_shock'][0]*100:.0f}% to {SCENARIO_CONFIG['severe']['utilities_shock'][1]*100:.0f}%",
                "Historical_Worst": f"{util_worst_60d*100:.1f}%" if not np.isnan(util_worst_60d) else "N/A",
                "Status": "✓ SUPPORTED" if util_worst_60d <= SCENARIO_CONFIG['severe']['utilities_shock'][0] else "⚠ AGGRESSIVE"
            },
            {
                "Scenario": "Moderate 10-day",
                "Asset_Class": "AI/Tech",
                "Proposed_Shock": f"{SCENARIO_CONFIG['moderate']['single_name_shock'][0]*100:.0f}% to {SCENARIO_CONFIG['moderate']['single_name_shock'][1]*100:.0f}%",
                "Historical_Worst": f"{ai_worst_10d*100:.1f}%" if not np.isnan(ai_worst_10d) else "N/A",
                "Status": "✓ SUPPORTED" if ai_worst_10d <= SCENARIO_CONFIG['moderate']['single_name_shock'][0] else "⚠ AGGRESSIVE"
            },
            {
                "Scenario": "Moderate 10-day",
                "Asset_Class": "Utilities",
                "Proposed_Shock": f"{SCENARIO_CONFIG['moderate']['utilities_shock'][0]*100:.0f}% to {SCENARIO_CONFIG['moderate']['utilities_shock'][1]*100:.0f}%",
                "Historical_Worst": f"{util_worst_10d*100:.1f}%" if not np.isnan(util_worst_10d) else "N/A",
                "Status": "✓ SUPPORTED" if util_worst_10d <= SCENARIO_CONFIG['moderate']['utilities_shock'][0] else "⚠ AGGRESSIVE"
            }
        ]
        
        validation_df = pd.DataFrame(validation)
        
        return {
            "betas": betas,
            "period_details": period_df,
            "period_summary": summary_df,
            "validation": validation_df
        }


# =============================================================================
# DATA FETCHING
# =============================================================================

def fetch_data_yfinance(tickers: List[str], start: str, end: str) -> pd.DataFrame:
    """Fetch data from Yahoo Finance"""
    if not YF_AVAILABLE:
        raise ImportError("yfinance not installed. Run: pip install yfinance")
    
    print(f"Fetching {len(tickers)} tickers from {start} to {end}...")
    data = yf.download(tickers, start=start, end=end, progress=False)
    
    if isinstance(data.columns, pd.MultiIndex):
        prices = data['Adj Close']
    else:
        prices = data
    
    return prices.dropna(how='all')


def load_data_csv(filepath: str) -> pd.DataFrame:
    """Load data from CSV file"""
    df = pd.read_csv(filepath, parse_dates=['Date'], index_col='Date')
    return df


def load_data_excel(filepath: str) -> pd.DataFrame:
    """Load data from Excel file"""
    df = pd.read_excel(filepath, parse_dates=['Date'], index_col='Date')
    return df


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_to_excel(results: Dict, filepath: str):
    """Export results to formatted Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for sheet_name, df in results.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    wb.save(filepath)
    print(f"Saved to: {filepath}")


# =============================================================================
# GUI APPLICATION
# =============================================================================

class StressAnalyzerGUI:
    """GUI for the Stress Analyzer"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("AI Bubble Stress Scenario Analyzer")
        self.root.geometry("900x700")
        
        self.prices = None
        self.results = None
        
        self.create_widgets()
    
    def create_widgets(self):
        """Create GUI widgets"""
        
        # Main notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Tab 1: Data Input
        data_frame = ttk.Frame(notebook)
        notebook.add(data_frame, text="1. Load Data")
        self.create_data_tab(data_frame)
        
        # Tab 2: Configuration
        config_frame = ttk.Frame(notebook)
        notebook.add(config_frame, text="2. Configure")
        self.create_config_tab(config_frame)
        
        # Tab 3: Results
        results_frame = ttk.Frame(notebook)
        notebook.add(results_frame, text="3. Results")
        self.create_results_tab(results_frame)
    
    def create_data_tab(self, parent):
        """Create data input tab"""
        
        # Yahoo Finance section
        yf_frame = ttk.LabelFrame(parent, text="Option 1: Fetch from Yahoo Finance")
        yf_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(yf_frame, text="Start Date (YYYY-MM-DD):").grid(row=0, column=0, padx=5, pady=5)
        self.start_entry = ttk.Entry(yf_frame, width=15)
        self.start_entry.insert(0, "1999-01-01")
        self.start_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(yf_frame, text="End Date:").grid(row=0, column=2, padx=5, pady=5)
        self.end_entry = ttk.Entry(yf_frame, width=15)
        self.end_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.end_entry.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Button(yf_frame, text="Fetch Data", command=self.fetch_yf_data).grid(row=0, column=4, padx=10, pady=5)
        
        # File upload section
        file_frame = ttk.LabelFrame(parent, text="Option 2: Load from File (CSV/Excel)")
        file_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(file_frame, text="File must have 'Date' column and ticker columns with Adj Close prices").pack(pady=5)
        ttk.Button(file_frame, text="Browse for CSV", command=lambda: self.load_file('csv')).pack(side='left', padx=10, pady=5)
        ttk.Button(file_frame, text="Browse for Excel", command=lambda: self.load_file('excel')).pack(side='left', padx=10, pady=5)
        
        # Paste data section
        paste_frame = ttk.LabelFrame(parent, text="Option 3: Paste CSV Data")
        paste_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        ttk.Label(paste_frame, text="Paste CSV (Date,SPY,QQQ,NVDA,...):").pack(anchor='w', padx=5)
        self.paste_text = scrolledtext.ScrolledText(paste_frame, height=10)
        self.paste_text.pack(fill='both', expand=True, padx=5, pady=5)
        ttk.Button(paste_frame, text="Load Pasted Data", command=self.load_pasted_data).pack(pady=5)
        
        # Status
        self.data_status = ttk.Label(parent, text="No data loaded", font=('Arial', 10, 'bold'))
        self.data_status.pack(pady=10)
    
    def create_config_tab(self, parent):
        """Create configuration tab"""
        
        # Tickers
        ticker_frame = ttk.LabelFrame(parent, text="Tickers to Analyze")
        ticker_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(ticker_frame, text="AI/Tech Tickers (comma-separated):").pack(anchor='w', padx=5, pady=2)
        self.ai_tickers_entry = ttk.Entry(ticker_frame, width=80)
        self.ai_tickers_entry.insert(0, ",".join(DEFAULT_TICKERS["ai_tech"]))
        self.ai_tickers_entry.pack(fill='x', padx=5, pady=2)
        
        ttk.Label(ticker_frame, text="Utility Tickers (comma-separated):").pack(anchor='w', padx=5, pady=2)
        self.util_tickers_entry = ttk.Entry(ticker_frame, width=80)
        self.util_tickers_entry.insert(0, ",".join(DEFAULT_TICKERS["utilities"]))
        self.util_tickers_entry.pack(fill='x', padx=5, pady=2)
        
        # Scenario parameters
        scenario_frame = ttk.LabelFrame(parent, text="Scenario Shock Parameters")
        scenario_frame.pack(fill='x', padx=10, pady=10)
        
        # Severe
        ttk.Label(scenario_frame, text="SEVERE (60-day):", font=('Arial', 10, 'bold')).grid(row=0, column=0, columnspan=4, pady=5)
        
        ttk.Label(scenario_frame, text="Single Names Low %:").grid(row=1, column=0, padx=5)
        self.severe_sn_low = ttk.Entry(scenario_frame, width=8)
        self.severe_sn_low.insert(0, "-50")
        self.severe_sn_low.grid(row=1, column=1, padx=5)
        
        ttk.Label(scenario_frame, text="High %:").grid(row=1, column=2, padx=5)
        self.severe_sn_high = ttk.Entry(scenario_frame, width=8)
        self.severe_sn_high.insert(0, "-40")
        self.severe_sn_high.grid(row=1, column=3, padx=5)
        
        ttk.Label(scenario_frame, text="Utilities Low %:").grid(row=2, column=0, padx=5)
        self.severe_util_low = ttk.Entry(scenario_frame, width=8)
        self.severe_util_low.insert(0, "-30")
        self.severe_util_low.grid(row=2, column=1, padx=5)
        
        ttk.Label(scenario_frame, text="High %:").grid(row=2, column=2, padx=5)
        self.severe_util_high = ttk.Entry(scenario_frame, width=8)
        self.severe_util_high.insert(0, "-15")
        self.severe_util_high.grid(row=2, column=3, padx=5)
        
        # Moderate
        ttk.Label(scenario_frame, text="MODERATE (10-day):", font=('Arial', 10, 'bold')).grid(row=3, column=0, columnspan=4, pady=5)
        
        ttk.Label(scenario_frame, text="Single Names Low %:").grid(row=4, column=0, padx=5)
        self.mod_sn_low = ttk.Entry(scenario_frame, width=8)
        self.mod_sn_low.insert(0, "-25")
        self.mod_sn_low.grid(row=4, column=1, padx=5)
        
        ttk.Label(scenario_frame, text="High %:").grid(row=4, column=2, padx=5)
        self.mod_sn_high = ttk.Entry(scenario_frame, width=8)
        self.mod_sn_high.insert(0, "-15")
        self.mod_sn_high.grid(row=4, column=3, padx=5)
        
        ttk.Label(scenario_frame, text="Utilities Low %:").grid(row=5, column=0, padx=5)
        self.mod_util_low = ttk.Entry(scenario_frame, width=8)
        self.mod_util_low.insert(0, "-15")
        self.mod_util_low.grid(row=5, column=1, padx=5)
        
        ttk.Label(scenario_frame, text="High %:").grid(row=5, column=2, padx=5)
        self.mod_util_high = ttk.Entry(scenario_frame, width=8)
        self.mod_util_high.insert(0, "-8")
        self.mod_util_high.grid(row=5, column=3, padx=5)
        
        # Run button
        ttk.Button(parent, text="▶ RUN ANALYSIS", command=self.run_analysis, 
                   style='Accent.TButton').pack(pady=20)
    
    def create_results_tab(self, parent):
        """Create results tab"""
        
        # Results text
        self.results_text = scrolledtext.ScrolledText(parent, height=25, font=('Courier', 10))
        self.results_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Export button
        export_frame = ttk.Frame(parent)
        export_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(export_frame, text="Export to Excel", command=self.export_results).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Copy to Clipboard", command=self.copy_results).pack(side='left', padx=5)
    
    def fetch_yf_data(self):
        """Fetch data from Yahoo Finance"""
        try:
            ai_tickers = [t.strip() for t in self.ai_tickers_entry.get().split(",")]
            util_tickers = [t.strip() for t in self.util_tickers_entry.get().split(",")]
            all_tickers = list(set(ai_tickers + util_tickers + ["SPY", "QQQ"]))
            
            self.prices = fetch_data_yfinance(
                all_tickers,
                self.start_entry.get(),
                self.end_entry.get()
            )
            
            self.data_status.config(
                text=f"✓ Loaded {len(self.prices)} days, {len(self.prices.columns)} tickers",
                foreground="green"
            )
            messagebox.showinfo("Success", f"Loaded {len(self.prices)} trading days")
        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    def load_file(self, file_type):
        """Load data from file"""
        try:
            if file_type == 'csv':
                filepath = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
                if filepath:
                    self.prices = load_data_csv(filepath)
            else:
                filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
                if filepath:
                    self.prices = load_data_excel(filepath)
            
            if self.prices is not None:
                self.data_status.config(
                    text=f"✓ Loaded {len(self.prices)} days, {len(self.prices.columns)} tickers",
                    foreground="green"
                )
        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    def load_pasted_data(self):
        """Load data from pasted CSV"""
        try:
            from io import StringIO
            csv_data = self.paste_text.get("1.0", tk.END)
            self.prices = pd.read_csv(StringIO(csv_data), parse_dates=['Date'], index_col='Date')
            
            self.data_status.config(
                text=f"✓ Loaded {len(self.prices)} days, {len(self.prices.columns)} tickers",
                foreground="green"
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    def run_analysis(self):
        """Run the analysis"""
        if self.prices is None:
            messagebox.showerror("Error", "Please load data first!")
            return
        
        try:
            # Update scenario config from GUI
            SCENARIO_CONFIG["severe"]["single_name_shock"] = (
                float(self.severe_sn_low.get()) / 100,
                float(self.severe_sn_high.get()) / 100
            )
            SCENARIO_CONFIG["severe"]["utilities_shock"] = (
                float(self.severe_util_low.get()) / 100,
                float(self.severe_util_high.get()) / 100
            )
            SCENARIO_CONFIG["moderate"]["single_name_shock"] = (
                float(self.mod_sn_low.get()) / 100,
                float(self.mod_sn_high.get()) / 100
            )
            SCENARIO_CONFIG["moderate"]["utilities_shock"] = (
                float(self.mod_util_low.get()) / 100,
                float(self.mod_util_high.get()) / 100
            )
            
            # Get tickers
            ai_tickers = [t.strip() for t in self.ai_tickers_entry.get().split(",")]
            util_tickers = [t.strip() for t in self.util_tickers_entry.get().split(",")]
            
            # Run analysis
            analyzer = StressAnalyzer(self.prices)
            self.results = analyzer.run_full_analysis(ai_tickers, util_tickers)
            
            # Display results
            self.display_results()
            
            messagebox.showinfo("Success", "Analysis complete! See Results tab.")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))
            import traceback
            traceback.print_exc()
    
    def display_results(self):
        """Display results in text area"""
        self.results_text.delete("1.0", tk.END)
        
        text = "=" * 70 + "\n"
        text += "AI BUBBLE STRESS SCENARIO - ANALYSIS RESULTS\n"
        text += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        text += "=" * 70 + "\n\n"
        
        # Validation Summary
        text += "SHOCK VALIDATION SUMMARY\n"
        text += "-" * 70 + "\n"
        text += self.results["validation"].to_string(index=False) + "\n\n"
        
        # Period Summary
        text += "HISTORICAL PERIOD SUMMARY\n"
        text += "-" * 70 + "\n"
        text += self.results["period_summary"].to_string(index=False) + "\n\n"
        
        # Beta Analysis
        text += "BETA ANALYSIS\n"
        text += "-" * 70 + "\n"
        text += self.results["betas"].to_string(index=False) + "\n\n"
        
        # Detailed Period Results
        text += "DETAILED PERIOD RESULTS\n"
        text += "-" * 70 + "\n"
        text += self.results["period_details"].to_string(index=False) + "\n"
        
        self.results_text.insert("1.0", text)
    
    def export_results(self):
        """Export results to Excel"""
        if self.results is None:
            messagebox.showerror("Error", "Run analysis first!")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfilename=f"stress_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if filepath:
            export_to_excel({
                "Validation": self.results["validation"],
                "Period_Summary": self.results["period_summary"],
                "Betas": self.results["betas"],
                "Period_Details": self.results["period_details"]
            }, filepath)
            messagebox.showinfo("Success", f"Exported to {filepath}")
    
    def copy_results(self):
        """Copy results to clipboard"""
        self.root.clipboard_clear()
        self.root.clipboard_append(self.results_text.get("1.0", tk.END))
        messagebox.showinfo("Success", "Results copied to clipboard!")
    
    def run(self):
        """Run the GUI"""
        self.root.mainloop()


# =============================================================================
# COMMAND LINE MODE (fallback if no tkinter)
# =============================================================================

def run_cli_mode():
    """Run in command line mode"""
    print("\n" + "=" * 60)
    print("AI BUBBLE STRESS SCENARIO ANALYZER - CLI MODE")
    print("=" * 60)
    
    print("\nData source options:")
    print("  1. Fetch from Yahoo Finance")
    print("  2. Load from CSV file")
    print("  3. Load from Excel file")
    
    choice = input("\nSelect option (1-3): ").strip()
    
    if choice == "1":
        if not YF_AVAILABLE:
            print("yfinance not installed. Run: pip install yfinance")
            return
        
        start = input("Start date (YYYY-MM-DD) [1999-01-01]: ").strip() or "1999-01-01"
        end = input("End date (YYYY-MM-DD) [today]: ").strip() or datetime.now().strftime("%Y-%m-%d")
        
        all_tickers = DEFAULT_TICKERS["ai_tech"] + DEFAULT_TICKERS["utilities"] + DEFAULT_TICKERS["benchmarks"]
        prices = fetch_data_yfinance(all_tickers, start, end)
        
    elif choice == "2":
        filepath = input("CSV file path: ").strip()
        prices = load_data_csv(filepath)
        
    elif choice == "3":
        filepath = input("Excel file path: ").strip()
        prices = load_data_excel(filepath)
    else:
        print("Invalid option")
        return
    
    print(f"\nLoaded {len(prices)} trading days, {len(prices.columns)} tickers")
    
    # Run analysis
    analyzer = StressAnalyzer(prices)
    results = analyzer.run_full_analysis(
        DEFAULT_TICKERS["ai_tech"],
        DEFAULT_TICKERS["utilities"]
    )
    
    # Print results
    print("\n" + "=" * 60)
    print("SHOCK VALIDATION SUMMARY")
    print("=" * 60)
    print(results["validation"].to_string(index=False))
    
    print("\n" + "=" * 60)
    print("PERIOD SUMMARY")
    print("=" * 60)
    print(results["period_summary"].to_string(index=False))
    
    print("\n" + "=" * 60)
    print("BETA ANALYSIS")
    print("=" * 60)
    print(results["betas"].to_string(index=False))
    
    # Export
    export = input("\nExport to Excel? (y/n): ").strip().lower()
    if export == 'y':
        filepath = input("Output path [stress_analysis.xlsx]: ").strip() or "stress_analysis.xlsx"
        export_to_excel({
            "Validation": results["validation"],
            "Period_Summary": results["period_summary"],
            "Betas": results["betas"],
            "Period_Details": results["period_details"]
        }, filepath)


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    if TK_AVAILABLE:
        app = StressAnalyzerGUI()
        app.run()
    else:
        run_cli_mode()
